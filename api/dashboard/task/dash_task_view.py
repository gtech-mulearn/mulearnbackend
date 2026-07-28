import uuid
from django.db.models import Count, Exists, OuterRef, Q

from rest_framework import status
from rest_framework.views import APIView

from db.organization import Organization, UserOrganizationLink
from db.task import Channel, InterestGroup, Level, TaskList, TaskType, UserIgLink
from db.user import UserMentor
from db.skill import Skill, TaskSkillLink
from utils.permission import CustomizePermission, JWTUtils, role_required
from utils.response import CustomResponse
from utils.types import Events, OrganizationType, RoleType
from utils.utils import CommonUtils, DateTimeUtils, ImportCSV
from .dash_task_serializer import (
    TaskImportSerializer,
    TaskListPublicSerializer,
    TaskListSerializer,
    TaskModifySerializer,
    TaskTypeCreateUpdateSerializer,
    TasktypeSerializer,
)

from openpyxl import load_workbook
from tempfile import NamedTemporaryFile
from io import BytesIO
from django.http import FileResponse
from drf_spectacular.utils import extend_schema, inline_serializer, OpenApiResponse, OpenApiParameter
from drf_spectacular.openapi import OpenApiTypes
from rest_framework import serializers as s
from utils.schema_utils import CustomResponseSerializer


class TaskPublicListAPI(APIView):
    authentication_classes = [CustomizePermission]

    @extend_schema(
        tags=['Dashboard - Task'],
        description=(
            "Retrieve active tasks grouped into three journey sections: "
            "'start_journey' (generic, level-ordered tasks — excludes IG, intern and event tasks), "
            "'become_expert' (the caller's IG task(s) plus company-submitted tasks, level-ordered), "
            "'events' (event-linked tasks visible to the caller)."
        ),
        parameters=[
            OpenApiParameter(
                "ig_id",
                OpenApiTypes.STR,
                OpenApiParameter.QUERY,
                required=False,
                description=(
                    "Override which Interest Group's tasks appear under become_expert. "
                    "Defaults to every IG the caller actively belongs to. Works for "
                    "unauthenticated callers too (any IG can be previewed)."
                ),
            ),
        ],
        responses={200: inline_serializer("TaskPublicListResponse", fields={
            "start_journey": s.ListField(child=s.DictField()),
            "become_expert": s.ListField(child=s.DictField()),
            "events": s.ListField(child=s.DictField()),
        })},
    )
    def get(self, request):
        # Local imports to avoid circular import issues
        from api.dashboard.events.public_views import _get_viewer_id, _build_scope_filter
        from api.dashboard.events.serializers import get_live_events
        from db.events import Event
        from db.user import UserRoleLink

        base_queryset = TaskList.objects.select_related(
            "channel", "type", "level", "ig", "org", "event_fk", "requested_by"
        ).filter(active=True)

        # --- Visibility filtering ---
        # Determine which orgs/IGs the current user belongs to (if authenticated).
        # Unauthenticated users can only see global tasks and company tasks.
        campus_org_types = [OrganizationType.COLLEGE.value, OrganizationType.SCHOOL.value]
        is_authenticated = JWTUtils.is_logged_in(request)

        if is_authenticated:
            user_id = JWTUtils.fetch_user_id(request)

            learner_org_ids = list(
                UserOrganizationLink.objects.filter(
                    user_id=user_id,
                    verified=True,
                    org__org_type__in=campus_org_types,
                ).values_list("org_id", flat=True)
            )

            # Org visibility rule:
            #   - No org set (global task)                → visible to all
            #   - Org is a Company                        → visible to all
            #   - Org is a Campus (College/School)        → only learners of that campus
            org_filter = (
                Q(org__isnull=True)
                | Q(org__org_type=OrganizationType.COMPANY.value)
                | Q(org__org_type__in=campus_org_types, org_id__in=learner_org_ids)
            )

            member_ig_ids = list(
                UserIgLink.objects.filter(
                    user_id=user_id,
                    is_active=True,
                ).values_list("ig_id", flat=True)
            )
        else:
            # Unauthenticated: global + company tasks only
            org_filter = Q(org__isnull=True) | Q(org__org_type=OrganizationType.COMPANY.value)
            member_ig_ids = []

        # ---------- start_journey: generic, level-ordered tasks ----------
        # No IG, no event, and not an "intern" task (same convention used by
        # get_karma_breakdown: hashtag prefix or an active Intern/Intern Lead creator role).
        is_intern_creator = Exists(
            UserRoleLink.objects.filter(
                user=OuterRef("requested_by"),
                role__title__in=[RoleType.INTERN.value, RoleType.INTERN_LEAD.value],
                is_active=True,
            )
        )
        start_journey_qs = (
            base_queryset.filter(event_fk__isnull=True, ig__isnull=True)
            .filter(org_filter)
            .annotate(is_intern_creator=is_intern_creator)
            .exclude(hashtag__startswith="#intern-")
            .exclude(is_intern_creator=True)
            .order_by("level__level_order", "title")
        )

        # ---------- become_expert & events: authenticated callers only ----------
        # Unauthenticated callers only get start_journey.
        become_expert_data = []
        events_data = []

        if is_authenticated:
            # become_expert: caller's IG task(s) + company tasks, level-ordered
            ig_id_param = request.query_params.get("ig_id")
            target_ig_ids = [ig_id_param] if ig_id_param else member_ig_ids

            become_expert_filter = Q(
                requested_by__isnull=False,
                requested_by__company_profile__isnull=False,
            )
            if target_ig_ids:
                become_expert_filter |= Q(ig_id__in=target_ig_ids)

            become_expert_qs = (
                base_queryset.filter(event_fk__isnull=True)
                .filter(become_expert_filter)
                .order_by("level__level_order", "title")
                .distinct()
            )
            become_expert_data = TaskListPublicSerializer(become_expert_qs, many=True).data

            # events: event-linked tasks visible to the caller
            viewer_id = _get_viewer_id(request)
            scope_filter = _build_scope_filter(viewer_id)
            accessible_event_ids = list(
                get_live_events()
                .filter(
                    scope_filter,
                    status__in=[Event.Status.PUBLISHED, Event.Status.ONGOING],
                )
                .values_list("id", flat=True)
            )
            events_qs = base_queryset.filter(
                event_fk__isnull=False,
                event_fk_id__in=accessible_event_ids,
            ).order_by("event_fk__title", "title")
            events_data = TaskListPublicSerializer(events_qs, many=True).data

        return CustomResponse(
            response={
                "start_journey": TaskListPublicSerializer(start_journey_qs, many=True).data,
                "become_expert": become_expert_data,
                "events": events_data,
            }
        ).get_success_response()


class TaskListAPI(APIView):
    authentication_classes = [CustomizePermission]

    @role_required(
        [
            RoleType.ADMIN.value,
            RoleType.FELLOW.value,
            RoleType.ASSOCIATE.value,
        ]
    )
    @extend_schema(
        tags=['Dashboard - Task'],
        description="Retrieve Task List.",
        parameters=[
            OpenApiParameter(
                "task_source",
                OpenApiTypes.STR,
                OpenApiParameter.QUERY,
                required=False,
                enum=["company", "ig_mentor", "campus_mentor", "platform"],
                description=(
                    "Filter tasks by creator type: "
                    "'company' = tasks submitted by a verified company user, "
                    "'ig_mentor' = tasks submitted by an approved IG mentor, "
                    "'campus_mentor' = tasks submitted by an approved campus mentor, "
                    "'platform' = tasks created by platform admins."
                ),
            ),
        ],
        responses={200: inline_serializer("TaskListResponse", fields={
            "data": s.ListField(child=s.DictField()),
            "pagination": s.DictField(),
        })},
    )
    def get(self, request):
        task_queryset = TaskList.objects.select_related(
            "created_by", "updated_by", "channel", "type", "level", "ig", "org", "requested_by"
        ).prefetch_related(
            "skill_links__skill"
        ).annotate(
            total_karma_gainers_count=Count("karma_activity_log_task", filter=Q(karma_activity_log_task__appraiser_approved=True))
        )

        task_source = request.query_params.get("task_source")
        if task_source == "company":
            task_queryset = task_queryset.filter(
                requested_by__isnull=False,
                requested_by__company_profile__isnull=False,
            )
        elif task_source == "ig_mentor":
            task_queryset = task_queryset.filter(
                requested_by__isnull=False,
                requested_by__user_mentor_user__mentor_tier=UserMentor.MentorTier.IG_MENTOR,
                requested_by__user_mentor_user__status=UserMentor.Status.APPROVED,
            ).distinct()
        elif task_source == "campus_mentor":
            task_queryset = task_queryset.filter(
                requested_by__isnull=False,
                requested_by__user_mentor_user__mentor_tier=UserMentor.MentorTier.CAMPUS_MENTOR,
                requested_by__user_mentor_user__status=UserMentor.Status.APPROVED,
            ).distinct()
        elif task_source == "platform":
            task_queryset = task_queryset.filter(requested_by__isnull=True)

        paginated_queryset = CommonUtils.get_paginated_queryset(
            task_queryset,
            request,
            search_fields=[
                "hashtag",
                "title",
                "description",
                "karma",
                "channel__name",
                "type__title",
                "active",
                "variable_karma",
                "usage_count",
                "level__name",
                "org__title",
                "ig__name",
                "event",
                "updated_at",
                "updated_by__full_name",
                "created_by__full_name",
                "created_at",
            ],
            sort_fields={
                "hashtag": "hashtag",
                "title": "title",
                "description": "description",
                "karma": "karma",
                "channels": "channel__name",
                "type": "type__title",
                "active": "active",
                "variable_karma": "variable_karma",
                "usage_count": "usage_count",
                "level": "level__name",
                "org": "org__title",
                "ig": "ig__name",
                "event": "event",
                "updated_at": "updated_at",
                "updated_by": "updated_by__full_name",
                "created_by": "created_by__full_name",
                "created_at": "created_at",
            },
        )

        task_serializer_data = TaskListSerializer(
            paginated_queryset.get("queryset"), many=True
        ).data

        return CustomResponse().paginated_response(
            data=task_serializer_data,
            pagination=paginated_queryset.get("pagination"),
        )

    @role_required(
        [
            RoleType.ADMIN.value,
            RoleType.FELLOW.value,
            RoleType.ASSOCIATE.value,
        ]
    )
    @extend_schema(
        tags=['Dashboard - Task'],
        description="Create Task List.",
        request=TaskModifySerializer,
        responses={200: OpenApiResponse(description="Task Created Successfully")},
    )
    def post(self, request):  # create
        user_id = JWTUtils.fetch_user_id(request)

        mutable_data = request.data.copy()  # Create a mutable copy of request.data
        mutable_data["created_by"] = user_id
        mutable_data["updated_by"] = user_id
        
        # Extract skill_ids before serializer processing
        skill_ids = mutable_data.pop("skill_ids", None)
        if isinstance(skill_ids, str):
            import json
            try:
                skill_ids = json.loads(skill_ids)
            except:
                skill_ids = []

        serializer = TaskModifySerializer(data=mutable_data, context={"request": request})

        if not serializer.is_valid():
            return CustomResponse(message=serializer.errors).get_failure_response()

        event_fk_id = serializer.validated_data.get("event_fk_id")
        if event_fk_id and not _is_event_accessible(request, event_fk_id):
            return CustomResponse(
                message={"event_id": ["Selected event does not exist or is not accessible."]}
            ).get_failure_response()

        task = serializer.save()
        
        # Handle skill links
        if skill_ids:
            self._save_task_skills(task.id, skill_ids, user_id)
        
        return CustomResponse(
            general_message="Task Created Successfully"
        ).get_success_response()
    
    def _save_task_skills(self, task_id, skill_ids, user_id):
        """Save skill links for a task"""
        # Clear existing links
        TaskSkillLink.objects.filter(task_id=task_id).delete()
        
        # Create new links
        for skill_id in skill_ids:
            if Skill.objects.filter(id=skill_id, is_active=True).exists():
                TaskSkillLink.objects.create(
                    id=str(uuid.uuid4()),
                    task_id=task_id,
                    skill_id=skill_id,
                    created_by_id=user_id,
                )


class TaskAPI(APIView):
    authentication_classes = [CustomizePermission]

    @role_required(
        [
            RoleType.ADMIN.value,
            RoleType.FELLOW.value,
            RoleType.ASSOCIATE.value,
        ]
    )
    @extend_schema(
        tags=['Dashboard - Task'],
        description="Retrieve Task.",
        responses={200: TaskModifySerializer},
    )
    def get(self, request, task_id):
        try:
            task_queryset = TaskList.objects.get(pk=task_id)
        except TaskList.DoesNotExist:
            return CustomResponse(
                general_message="Task not found."
            ).get_failure_response(status_code=404, http_status_code=status.HTTP_404_NOT_FOUND)

        task_serializer = TaskModifySerializer(task_queryset, many=False)
        return CustomResponse(response=task_serializer.data).get_success_response()

    @role_required(
        [
            RoleType.ADMIN.value,
            RoleType.FELLOW.value,
            RoleType.ASSOCIATE.value,
        ]
    )
    @extend_schema(tags=['Dashboard - Task'], description="Update Task.",
        responses={200: TaskModifySerializer},
    )
    def put(self, request, task_id):  # edit

        user_id = JWTUtils.fetch_user_id(request)
        mutable_data = request.data.copy()  # Create a mutable copy of request.data
        mutable_data["updated_by"] = user_id
        
        # Extract skill_ids before serializer processing
        skill_ids = mutable_data.pop("skill_ids", None)
        if isinstance(skill_ids, str):
            import json
            try:
                skill_ids = json.loads(skill_ids)
            except:
                skill_ids = None

        try:
            task = TaskList.objects.get(pk=task_id)
        except TaskList.DoesNotExist:
            return CustomResponse(
                general_message="Task not found."
            ).get_failure_response(status_code=404, http_status_code=status.HTTP_404_NOT_FOUND)

        serializer = TaskModifySerializer(
            task, data=mutable_data, partial=True, context={"request": request}
        )

        if not serializer.is_valid():
            return CustomResponse(message=serializer.errors).get_failure_response()

        if "event_fk_id" in serializer.validated_data:
            event_fk_id = serializer.validated_data.get("event_fk_id")
            if event_fk_id and not _is_event_accessible(request, event_fk_id):
                return CustomResponse(
                    message={"event_id": ["Selected event does not exist or is not accessible."]}
                ).get_failure_response()

        serializer.save()

        # Handle skill links if provided
        if skill_ids is not None:
            self._save_task_skills(task_id, skill_ids, user_id)

        return CustomResponse(general_message=serializer.data).get_success_response()
    
    def _save_task_skills(self, task_id, skill_ids, user_id):
        """Save skill links for a task"""
        # Clear existing links
        TaskSkillLink.objects.filter(task_id=task_id).delete()
        
        # Create new links
        for skill_id in skill_ids:
            if Skill.objects.filter(id=skill_id, is_active=True).exists():
                TaskSkillLink.objects.create(
                    id=str(uuid.uuid4()),
                    task_id=task_id,
                    skill_id=skill_id,
                    created_by_id=user_id,
                )

    @role_required(
        [
            RoleType.ADMIN.value,
            RoleType.FELLOW.value,
            RoleType.ASSOCIATE.value,
        ]
    )
    @extend_schema(tags=['Dashboard - Task'], description="Delete Task.",
        responses={200: TaskModifySerializer},
    )
    def delete(self, request, task_id):  # delete
        try:
            task = TaskList.objects.get(id=task_id)
        except TaskList.DoesNotExist:
            return CustomResponse(
                general_message="Task not found."
            ).get_failure_response(status_code=404, http_status_code=status.HTTP_404_NOT_FOUND)

        task.delete()

        return CustomResponse(
            general_message="Task deleted successfully"
        ).get_success_response()


class TaskListCSV(APIView):
    authentication_classes = [CustomizePermission]

    @role_required(
        [
            RoleType.ADMIN.value,
            RoleType.FELLOW.value,
            RoleType.ASSOCIATE.value,
        ]
    )
    @extend_schema(
        tags=['Dashboard - Task'],
        description="Retrieve Task List C S V.",
        responses={200: OpenApiResponse(description="CSV file download of the task list")},
    )
    def get(self, request):
        task_queryset = TaskList.objects.select_related(
            "created_by", "updated_by", "channel", "type", "level", "ig", "org"
        ).prefetch_related(
            "skill_links__skill"
        ).annotate(
            total_karma_gainers_count=Count("karma_activity_log_task", filter=Q(karma_activity_log_task__appraiser_approved=True))
        ).filter(active=True)

        task_queryset = CommonUtils.get_paginated_queryset(
            task_queryset,
            request,
            search_fields=[
                "hashtag",
                "title",
                "description",
                "karma",
                "channel__name",
                "type__title",
                "active",
                "variable_karma",
                "usage_count",
                "level__name",
                "org__title",
                "ig__name",
                "event",
                "updated_at",
                "updated_by__full_name",
                "created_by__full_name",
                "created_at",
            ],
            sort_fields={
                "hashtag": "hashtag",
                "title": "title",
                "description": "description",
                "karma": "karma",
                "channels": "channel__name",
                "type": "type__title",
                "active": "active",
                "variable_karma": "variable_karma",
                "usage_count": "usage_count",
                "level": "level__name",
                "org": "org__title",
                "ig": "ig__name",
                "event": "event",
                "updated_at": "updated_at",
                "updated_by": "updated_by__full_name",
                "created_by": "created_by__full_name",
                "created_at": "created_at",
            },
            is_pagination=False,
        )

        task_serializer_data = TaskListSerializer(task_queryset, many=True).data

        return CommonUtils.generate_csv(task_serializer_data, "Task List")


class ImportTaskListCSV(APIView):
    authentication_classes = [CustomizePermission]

    @role_required(
        [
            RoleType.ADMIN.value,
            RoleType.FELLOW.value,
            RoleType.ASSOCIATE.value,
        ]
    )
    @extend_schema(
        tags=['Dashboard - Task'],
        description="Create Import Task List C S V.",
        request=TaskImportSerializer,
        responses={200: TaskImportSerializer},
    )
    def post(self, request):
        file_obj = request.FILES.get("task_list", request.FILES.get("file"))
        if not file_obj:
            return CustomResponse(
                general_message="File not found."
            ).get_failure_response()

        excel_data = ImportCSV()
        excel_data = excel_data.read_excel_file(file_obj)

        if not excel_data:
            return CustomResponse(
                general_message="Empty csv file."
            ).get_failure_response()

        temp_headers = [
            "hashtag",
            "title",
            "description",
            "karma",
            "usage_count",
            "variable_karma",
            "level",
            "channel",
            "type",
            "ig",
            "org",
            "event",
        ]
        first_entry = excel_data[0]
        for key in temp_headers:
            if key not in first_entry:
                return CustomResponse(
                    general_message=f"{key} does not exist in the file."
                ).get_failure_response()

        excel_data = [row for row in excel_data if any(row.values())]
        valid_rows = []
        error_rows = []
        rows_to_validate = []

        hashtags_excel = set()
        hashtags_db = TaskList.objects.values_list("hashtag", flat=True)
        channels_to_fetch = set()
        task_types_to_fetch = set()
        levels_to_fetch = set()
        igs_to_fetch = set()
        orgs_to_fetch = set()

        for row in excel_data[1:]:
            hashtag = row.get("hashtag")
            if not hashtag:
                row["error"] = "Missing hashtag."
                error_rows.append(row)
                continue
            elif hashtag in hashtags_excel:
                row["error"] = f"Duplicate hashtag in excel: {hashtag}"
                error_rows.append(row)
                continue
            elif hashtag in hashtags_db:
                row["error"] = f"Duplicate hashtag in database: {hashtag}"
                error_rows.append(row)
                continue
            else:
                hashtags_excel.add(hashtag)

            title = row.get("title")
            if not title:
                row["error"] = "Missing title."
                error_rows.append(row)
                continue

            level = row.get("level")
            channel = row.get("channel")
            task_type = row.get("type")
            ig = row.get("ig")
            org = row.get("org")

            channels_to_fetch.add(channel)
            task_types_to_fetch.add(task_type)
            levels_to_fetch.add(level)
            igs_to_fetch.add(ig)
            orgs_to_fetch.add(org)
            rows_to_validate.append(row)

        channels = Channel.objects.filter(name__in=channels_to_fetch).values(
            "id", "name"
        )

        task_types = TaskType.objects.filter(title__in=task_types_to_fetch).values(
            "id", "title"
        )

        levels = Level.objects.filter(name__in=levels_to_fetch).values("id", "name")

        igs = InterestGroup.objects.filter(name__in=igs_to_fetch).values("id", "name")

        orgs = Organization.objects.filter(code__in=orgs_to_fetch).values("id", "code")

        channels_dict = {channel["name"]: channel["id"] for channel in channels}
        task_types_dict = {
            task_type["title"]: task_type["id"] for task_type in task_types
        }
        levels_dict = {level["name"]: level["id"] for level in levels}
        igs_dict = {ig["name"]: ig["id"] for ig in igs}
        orgs_dict = {org["code"]: org["id"] for org in orgs}
        events = Events.get_all_values()

        for row in rows_to_validate:
            level = row.pop("level")
            channel = row.pop("channel")
            task_type = row.pop("type")
            ig = row.pop("ig")
            org = row.pop("org")

            task_type_id = task_types_dict.get(task_type)
            channel_id = channels_dict.get(channel) if channel is not None else None
            level_id = levels_dict.get(level) if level is not None else None
            ig_id = igs_dict.get(ig) if ig is not None else None
            org_id = orgs_dict.get(org) if org is not None else None
            event = row.get("event")

            if channel and not channel_id:
                row["error"] = f"Invalid channel: {channel}"
                error_rows.append(row)
            elif not task_type_id:
                row["error"] = f"Invalid task type: {task_type}"
                error_rows.append(row)
            elif level and not level_id:
                row["error"] = f"Invalid level: {level}"
                error_rows.append(row)
            elif ig and not ig_id:
                row["error"] = f"Invalid interest group: {ig}"
                error_rows.append(row)
            elif org and not org_id:
                row["error"] = f"Invalid organization: {org}"
                error_rows.append(row)
            elif event is not None and event not in events:
                row["error"] = f"Invalid event: {event}"
                error_rows.append(row)
            else:
                user_id = JWTUtils.fetch_user_id(request)
                row["id"] = str(uuid.uuid4())
                row["updated_by_id"] = user_id
                row["updated_at"] = DateTimeUtils.get_current_utc_time()
                row["created_by_id"] = user_id
                row["created_at"] = DateTimeUtils.get_current_utc_time()
                row["active"] = True
                row["channel_id"] = channel_id or None
                row["type_id"] = task_type_id
                row["level_id"] = level_id or None
                row["ig_id"] = ig_id or None
                row["org_id"] = org_id or None
                
                valid_fields = [
                    "id", "hashtag", "discord_link", "title", "description", 
                    "karma", "channel_id", "type_id", "org_id", "event", "level_id", "ig_id", 
                    "active", "variable_karma", "usage_count", "created_by_id", 
                    "updated_by_id", "created_at", "updated_at"
                ]
                clean_row = {k: v for k, v in row.items() if k in valid_fields}
                valid_rows.append(clean_row)

        task_list_serializer = TaskImportSerializer(data=valid_rows, many=True)
        success_data = []
        if task_list_serializer.is_valid():
            task_list_serializer.save()
            for task_data in task_list_serializer.data:
                success_data.append(
                    {
                        "hashtag": task_data.get("hashtag", ""),
                        "title": task_data.get("title", ""),
                        "description": task_data.get("description", ""),
                        "karma": task_data.get("karma", ""),
                        "usage_count": task_data.get("usage_count", ""),
                        "variable_karma": task_data.get("variable_karma", ""),
                        "level": task_data.get("level_id", ""),
                        "channel": task_data.get("channel_id", ""),
                        "type": task_data.get("type_id", ""),
                        "ig": task_data.get("ig_id", ""),
                        "org": task_data.get("org_id", ""),
                        "event": task_data.get("event", ""),
                    }
                )
        else:
            error_rows.append(task_list_serializer.errors)

        return CustomResponse(
            response={"Success": success_data, "Failed": error_rows}
        ).get_success_response()


class ChannelDropdownAPI(APIView):
    authentication_classes = [CustomizePermission]

    @role_required(
        [
            RoleType.ADMIN.value,
            RoleType.FELLOW.value,
            RoleType.ASSOCIATE.value,
            RoleType.COMPANY.value,
            RoleType.MENTOR.value,
        ]
    )
    @extend_schema(tags=['Dashboard - Task'], description="Retrieve Channel Dropdown.",
        responses={200: inline_serializer("TaskChannelDropdownResponse", fields={
            "id": s.CharField(),
            "name": s.CharField(),
        })},
    )
    def get(self, request):
        channels = Channel.objects.values("id", "name")

        return CustomResponse(response=channels).get_success_response()


class IGDropdownAPI(APIView):
    authentication_classes = [CustomizePermission]

    @role_required(
        [
            RoleType.ADMIN.value,
            RoleType.FELLOW.value,
            RoleType.ASSOCIATE.value,
            RoleType.COMPANY.value,
            RoleType.MENTOR.value,
        ]
    )
    @extend_schema(tags=['Dashboard - Task'], description="Retrieve I G Dropdown.",
        responses={200: inline_serializer("TaskIGDropdownResponse", fields={
            "id": s.CharField(),
            "name": s.CharField(),
        })},
    )
    def get(self, request):
        igs = InterestGroup.objects.values("id", "name")
        return CustomResponse(response=igs).get_success_response()


class OrganizationDropdownAPI(APIView):
    authentication_classes = [CustomizePermission]

    @role_required(
        [
            RoleType.ADMIN.value,
            RoleType.FELLOW.value,
            RoleType.ASSOCIATE.value,
            RoleType.COMPANY.value,
            RoleType.MENTOR.value,
        ]
    )
    @extend_schema(tags=['Dashboard - Task'], description="Retrieve Organization Dropdown.",
        responses={200: inline_serializer("TaskOrganizationDropdownResponse", fields={
            "id": s.CharField(),
            "title": s.CharField(),
        })},
    )
    def get(self, request):
        organizations = Organization.objects.values("id", "title")
        return CustomResponse(response=organizations).get_success_response()


class LevelDropdownAPI(APIView):
    authentication_classes = [CustomizePermission]

    @role_required(
        [
            RoleType.ADMIN.value,
            RoleType.FELLOW.value,
            RoleType.ASSOCIATE.value,
            RoleType.COMPANY.value,
            RoleType.MENTOR.value,
        ]
    )
    @extend_schema(tags=['Dashboard - Task'], description="Retrieve Level Dropdown.",
        responses={200: inline_serializer("TaskLevelDropdownResponse", fields={
            "id": s.CharField(),
            "name": s.CharField(),
            "level_order": s.IntegerField(),
        })},
    )
    def get(self, request):
        # level_order is the numeric rank used by downstream consumers (e.g. job
        # eligibility "Min/Max Level" rules, which compare against a user's
        # level_order). Returning it lets clients store the order rather than an
        # opaque id/name.
        levels = Level.objects.values("id", "name", "level_order").order_by("level_order")
        return CustomResponse(response=levels).get_success_response()


class TaskTypesDropDownAPI(APIView):
    authentication_classes = [CustomizePermission]

    @role_required(
        [
            RoleType.ADMIN.value,
            RoleType.FELLOW.value,
            RoleType.ASSOCIATE.value,
            RoleType.COMPANY.value,
            RoleType.MENTOR.value,
        ]
    )
    @extend_schema(tags=['Dashboard - Task'], description="Retrieve Task Types Drop Down.",
        responses={200: TasktypeSerializer},
    )
    def get(self, request):
        task_types = TaskType.objects.values("id", "title")
        return CustomResponse(response=task_types).get_success_response()


def _is_event_accessible(request, event_id):
    """
    Authorization check: is this event live (PUBLISHED/ONGOING) and within
    the caller's visibility scope? Lives in the view layer, not the
    serializer, since this is a permission decision, not a data-integrity
    one. Reuses the same predicate accessible_event_ids/linkable-events
    trust, so a task can never link to an event the caller can't see.
    """
    from api.dashboard.events.public_views import _get_viewer_id, _build_scope_filter
    from api.dashboard.events.serializers import get_live_events
    from db.events import Event

    viewer_id = _get_viewer_id(request)
    scope_filter = _build_scope_filter(viewer_id)

    return get_live_events().filter(
        scope_filter,
        status__in=[Event.Status.PUBLISHED, Event.Status.ONGOING],
        id=event_id,
    ).exists()


class EventDropDownApi(APIView):
    authentication_classes = [CustomizePermission]

    @role_required(
        [
            RoleType.ADMIN.value,
        ]
    )
    @extend_schema(tags=['Dashboard - Task'], description="Retrieve Event Drop Down Api.",
        responses={200: inline_serializer("TaskEventDropdownResponse", fields={
            "events": s.ListField(child=s.CharField()),
        })},
    )
    def get(self, request):
        events = Events.get_all_values()
        return CustomResponse(response=events).get_success_response()


class TaskBaseTemplateAPI(APIView):
    authentication_classes = [CustomizePermission]

    @extend_schema(tags=['Dashboard - Task'], description="Retrieve Task Base Template.",
        responses={200: OpenApiResponse(description="Excel template file download (task_base_template.xlsx)")},
    )
    def get(self, request):
        wb = load_workbook("./excel-templates/task_base_template.xlsx")
        ws = wb["Data Definitions"]
        levels = Level.objects.all().values_list("name", flat=True)
        channels = Channel.objects.all().values_list("name", flat=True)
        task_types = TaskType.objects.all().values_list("title", flat=True)
        igs = InterestGroup.objects.all().values_list("name", flat=True)
        orgs = Organization.objects.all().values_list("code", flat=True)
        events = Events.get_all_values()

        data = {
            "level": levels,
            "channel": channels,
            "type": task_types,
            "ig": igs,
            "org": orgs,
            "event": events,
        }
        # Write data column-wise
        for col_num, (col_name, col_values) in enumerate(data.items(), start=1):
            for row, value in enumerate(col_values, start=2):
                ws.cell(row=row, column=col_num, value=value)
        # Save the file
        with NamedTemporaryFile() as tmp:
            tmp.close()  # with statement opened tmp, close it so wb.save can open it
            wb.save(tmp.name)
            with open(tmp.name, "rb") as f:
                f.seek(0)
                new_file_object = f.read()
        return FileResponse(
            BytesIO(new_file_object),
            as_attachment=True,
            filename="task_base_template.xlsx",
        )


class TaskTypeCrudAPI(APIView):
    authentication_classes = [CustomizePermission]

    @role_required(
        [
            RoleType.ADMIN.value,
            RoleType.COMPANY.value,
            RoleType.MENTOR.value,
        ]
    )
    @extend_schema(
        tags=['Dashboard - Task'],
        description="Retrieve Task Type Crud.",
        responses={200: TasktypeSerializer},
    )
    def get(self, request):
        taskType = TaskType.objects.all()
        paginated_queryset = CommonUtils.get_paginated_queryset(
            taskType,
            request,
            ["title"],
            {
                "title": "title",
                "updated_by": "updated_by",
                "created_by": "created_by",
                "updated_at": "updated_at",
                "created_at": "created_at",
            },
        )
        serializer = TasktypeSerializer(paginated_queryset.get("queryset"), many=True)

        return CustomResponse().paginated_response(
            data=serializer.data, pagination=paginated_queryset.get("pagination")
        )

    @role_required([RoleType.ADMIN.value])
    @extend_schema(
        tags=['Dashboard - Task'],
        description="Create Task Type Crud.",
        request=TaskTypeCreateUpdateSerializer,
        responses={200: TasktypeSerializer},
    )
    def post(self, request):
        user_id = JWTUtils.fetch_user_id(request)
        serializer = TaskTypeCreateUpdateSerializer(
            data=request.data, context={"user_id": user_id}
        )
        if serializer.is_valid():
            serializer.save()
            return CustomResponse(
                general_message="Task type added successfully"
            ).get_success_response()

        return CustomResponse(general_message=serializer.errors).get_failure_response()

    @role_required([RoleType.ADMIN.value])
    @extend_schema(tags=['Dashboard - Task'], description="Delete Task Type Crud.",
        responses={200: TasktypeSerializer},
    )
    def delete(self, request, task_type_id):
        taskType = TaskType.objects.filter(id=task_type_id).first()
        if taskType is None:
            return CustomResponse(
                general_message="task type doesnt exist"
            ).get_failure_response()
        taskType.delete()
        return CustomResponse(
            general_message=f"{taskType.title} Deleted Successfully"
        ).get_success_response()

    @role_required([RoleType.ADMIN.value])
    @extend_schema(
        tags=['Dashboard - Task'],
        description="Update Task Type Crud.",
        responses={200: TaskTypeCreateUpdateSerializer},
    )
    def put(self, request, task_type_id):
        taskType = TaskType.objects.filter(id=task_type_id).first()
        if taskType is None:
            return CustomResponse(
                general_message="task type not found"
            ).get_failure_response()
        serializer = TaskTypeCreateUpdateSerializer(
            taskType, data=request.data, context={"request": request}
        )
        if serializer.is_valid():
            serializer.save()
            return CustomResponse(
                general_message=f"{taskType.title} updated successfully"
            ).get_success_response()
        return CustomResponse(response=serializer.errors).get_failure_response()


# ---------------------------------------------------------------------------
# Admin: Task Approval Workflow (for company-submitted tasks)
# ---------------------------------------------------------------------------

class AdminTaskApprovalAPI(APIView):
    """
    GET  /dashboard/task/pending/              — list all tasks awaiting admin review
    PATCH /dashboard/task/<task_id>/approve/   — approve a pending task (goes live)
    PATCH /dashboard/task/<task_id>/reject/    — reject with a reason

    All actions require the Admin role.
    """
    authentication_classes = [CustomizePermission]

    @role_required([RoleType.ADMIN.value])
    @extend_schema(tags=['Dashboard - Task'], description="Retrieve Admin Task Approval.",
        responses={200: inline_serializer("TaskAdminApprovalListResponse", fields={
            "tasks": s.ListField(child=s.DictField()),
            "pagination": s.DictField(),
        })},
    )
    def get(self, request):
        """List tasks with filtering."""
        from django.utils import timezone as tz

        queryset = (
            TaskList.objects
            .select_related("ig", "type", "requested_by", "requested_by__company_profile")
            .order_by("created_at")
        )

        approval_status = request.query_params.get("approval_status", "pending")
        if approval_status:
            queryset = queryset.filter(approval_status=approval_status)

        source = request.query_params.get("source")
        role = request.query_params.get("role") or source
        
        if role == "mentor":
            queryset = queryset.filter(
                requested_by__isnull=False,
                requested_by__user_role_link_user__role__title=RoleType.MENTOR.value
            )
        elif role == "company":
            queryset = queryset.filter(
                requested_by__isnull=False,
                requested_by__user_role_link_user__role__title=RoleType.COMPANY.value
            )
        elif role == "admin":
            queryset = queryset.filter(requested_by__isnull=True)

        company_name = request.query_params.get("company_name")
        if company_name:
            queryset = queryset.filter(requested_by__company_profile__name__icontains=company_name)

        mentor_name = request.query_params.get("mentor_name")
        if mentor_name:
            queryset = queryset.filter(
                requested_by__full_name__icontains=mentor_name,
                requested_by__user_role_link_user__role__title=RoleType.MENTOR.value
            )

        paginated = CommonUtils.get_paginated_queryset(
            queryset,
            request,
            search_fields=[
                "title", "hashtag",
                "requested_by__company_profile__name",
                "requested_by__full_name",
            ],
            sort_fields={"createdAt": "created_at", "title": "title"},
            is_pagination=True,
        )

        data = []
        for task in paginated["queryset"]:
            company_profile = getattr(task.requested_by, "company_profile", None) if task.requested_by else None
            data.append({
                "id":                   str(task.id),
                "title":                task.title,
                "hashtag":              task.hashtag,
                "description":          task.description,
                "karma":                task.karma,
                "approval_status":      task.approval_status,
                "ig":                   {"id": str(task.ig.id), "name": task.ig.name} if task.ig else None,
                "type":                 {"id": str(task.type.id), "title": task.type.title} if task.type else None,
                "company_name":         company_profile.name if company_profile else None,
                "requested_by": {
                    "id":        str(task.requested_by.id),
                    "full_name": task.requested_by.full_name,
                } if task.requested_by else None,
                "requested_at": task.requested_at.isoformat() if task.requested_at else None,
                "created_at":   task.created_at.isoformat(),
            })

        return CustomResponse(
            general_message="Tasks fetched successfully.",
            response={"tasks": data, "pagination": paginated["pagination"]},
        ).get_success_response()

    @role_required([RoleType.ADMIN.value])
    @extend_schema(tags=['Dashboard - Task'], description="Partially update Admin Task Approval.",
        responses={200: inline_serializer("TaskAdminApprovalActionResponse", fields={
            "task_id": s.CharField(),
            "approval_status": s.CharField(),
            "active": s.BooleanField(),
            "rejection_reason": s.CharField(allow_null=True),
            "reviewed_by": s.CharField(allow_null=True),
            "reviewed_at": s.CharField(allow_null=True),
        })},
    )
    def patch(self, request, task_id):
        """
        Approve or reject a pending task.
        Body: { "action": "approve" | "reject", "reason": "<string — required for reject>" }
        """
        from django.utils import timezone as tz

        action = request.data.get("action")
        if action not in ("approve", "reject"):
            return CustomResponse(
                general_message="Invalid action. Must be 'approve' or 'reject'.",
                message={"error_code": "INVALID_ACTION"},
            ).get_failure_response()

        try:
            task = TaskList.objects.get(id=task_id)
        except TaskList.DoesNotExist:
            return CustomResponse(
                general_message="Task not found.",
                message={"error_code": "TASK_NOT_FOUND"},
            ).get_failure_response()

        if task.approval_status != "pending":
            return CustomResponse(
                general_message=f"Only pending tasks can be reviewed. Current status: '{task.approval_status}'.",
                message={"error_code": "INVALID_STATUS_TRANSITION"},
            ).get_failure_response()

        admin_user_id = JWTUtils.fetch_user_id(request)
        from db.user import User as UserModel
        admin_user = UserModel.objects.filter(id=admin_user_id).first()

        now = tz.now()

        if action == "approve":
            task.approval_status = "approved"
            task.active = True
            task.rejection_reason = None
            task.reviewed_by_admin = admin_user
            task.reviewed_at = now
            task.updated_by = admin_user
            task.save(update_fields=[
                "approval_status", "active", "rejection_reason",
                "reviewed_by_admin", "reviewed_at", "updated_by", "updated_at",
            ])
            message = "Task approved and is now live."
        else:
            reason = (request.data.get("reason") or "").strip()
            if not reason:
                return CustomResponse(
                    general_message="A rejection reason is required.",
                    message={"error_code": "REASON_REQUIRED"},
                ).get_failure_response()
            task.approval_status = "rejected"
            task.active = False
            task.rejection_reason = reason
            task.reviewed_by_admin = admin_user
            task.reviewed_at = now
            task.updated_by = admin_user
            task.save(update_fields=[
                "approval_status", "active", "rejection_reason",
                "reviewed_by_admin", "reviewed_at", "updated_by", "updated_at",
            ])
            message = "Task rejected."

        return CustomResponse(
            general_message=message,
            response={
                "task_id":          str(task.id),
                "approval_status":  task.approval_status,
                "active":           task.active,
                "rejection_reason": task.rejection_reason,
                "reviewed_by":      str(admin_user.id) if admin_user else None,
                "reviewed_at":      task.reviewed_at.isoformat() if task.reviewed_at else None,
            },
        ).get_success_response()
