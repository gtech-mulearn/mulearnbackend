import uuid
from email.mime.image import MIMEImage
from io import BytesIO
from tempfile import NamedTemporaryFile

import decouple
from django.core.mail import EmailMessage
from django.db import transaction
from django.db.models import Value
from django.db.models.functions import Coalesce
from django.http import FileResponse
from openpyxl import load_workbook
from rest_framework.views import APIView

from db.task import VoucherLog, TaskList
from db.user import User
from utils.karma_voucher import generate_karma_voucher, generate_ordered_id
from utils.permission import CustomizePermission, JWTUtils, role_required
from utils.response import CustomResponse
from utils.types import RoleType
from utils.utils import DateTimeUtils
from utils.utils import ImportCSV, CommonUtils
from .karma_voucher_serializer import VoucherLogCSVSerializer, VoucherLogSerializer, VoucherLogCreateSerializer, \
    VoucherLogUpdateSerializer


class ImportVoucherLogAPI(APIView):
    authentication_classes = [CustomizePermission]

    @role_required([RoleType.ADMIN.value, RoleType.FELLOW.value, RoleType.ASSOCIATE.value])
    def post(self, request):
        try:
            file_obj = request.FILES['voucher_log']
        except KeyError:
            return CustomResponse(general_message={'File not found.'}).get_failure_response()
        excel_data = ImportCSV()
        excel_data = excel_data.read_excel_file(file_obj)
        if not excel_data:
            return CustomResponse(general_message={'Empty csv file.'}).get_failure_response()

        temp_headers = ['muid', 'karma', 'hashtag', 'month', 'week']
        first_entry = excel_data[0]
        for key in temp_headers:
            if key not in first_entry:
                return CustomResponse(general_message={f'{key} does not exist in the file.'}).get_failure_response()
        current_user = JWTUtils.fetch_user_id(request)

        excel_data = [row for row in excel_data if any(row.values())]
        valid_rows = []
        error_rows = []
        success_rows = []
        users_to_fetch = set()
        tasks_to_fetch = set()
        for row in excel_data[1:]:
            task_hashtag = row.get('hashtag')
            muid = row.get('muid')
            users_to_fetch.add(muid)
            tasks_to_fetch.add(task_hashtag)
        # Fetching users and tasks in bulk
        users = User.objects.filter(muid__in=users_to_fetch).values('id', 'email', 'full_name', 'muid')
        tasks = TaskList.objects.filter(hashtag__in=tasks_to_fetch).values('id', 'hashtag')
        user_dict = {
            user['muid']: (
                user['id'], user['email'],
                user['full_name']
            ) for user in users
        }

        task_dict = {task['hashtag']: task['id'] for task in tasks}

        count = 1
        for row in excel_data[1:]:
            task_hashtag = row.get('hashtag')
            karma = row.get('karma')
            month = row.get('month')
            week = row.get('week')
            muid = row.get('muid')
            description = row.get('description')
            event = row.get('event')
            user_info = user_dict.get(muid)
            if user_info is None:
                row['error'] = f"Invalid muid: {muid}"
                error_rows.append(row)
            else:
                user_id, email, full_name = user_info
                task_id = task_dict.get(task_hashtag)
                if task_id is None:
                    row['error'] = f"Invalid task hashtag: {task_hashtag}"
                    error_rows.append(row)
                elif karma == 0:
                    row['error'] = "Karma cannot be 0"
                    error_rows.append(row)
                elif month is None:
                    row['error'] = "Month cannot be empty"
                    error_rows.append(row)
                else:
                    existing_codes = set(VoucherLog.objects.values_list('code', flat=True))
                    while generate_ordered_id(count) in existing_codes:
                        count += 1
                    # Preparing valid row data
                    row['user_id'] = user_id
                    row['task_id'] = task_id
                    row['id'] = str(uuid.uuid4())
                    row['code'] = generate_ordered_id(count)
                    row['claimed'] = False
                    row['created_by_id'] = current_user
                    row['updated_by_id'] = current_user
                    row['created_at'] = DateTimeUtils.get_current_utc_time()
                    row['updated_at'] = DateTimeUtils.get_current_utc_time()
                    count += 1
                    valid_rows.append(row)

        # Serializing and saving valid voucher rows to the database
        voucher_serializer = VoucherLogCSVSerializer(data=valid_rows, many=True)
        with transaction.atomic():
            if voucher_serializer.is_valid():
                    voucher_serializer.save()
            else:
                code_error_dict = {}
                for error in voucher_serializer.errors:
                    code_error = error.get('code')
                    error_msg = error.get('error')

                    code = str(code_error[0])
                    error_value = str(error_msg[0])
                    code_error_dict[code] = error_value

                for row in valid_rows:
                    code = row['code']
                    error_row = {}
                    if code in code_error_dict:
                        error_row['muid'] = row['muid']
                        error_row['karma'] = row['karma']
                        error_row['week'] = row['week']
                        error_row['month'] = row['month']
                        error_row['hashtag'] = row['hashtag']
                        error_row['description'] = row['description']
                        error_row['event'] = row['event']
                        error_row['error'] = code_error_dict[code]
                        error_rows.append(error_row)
                    
                return CustomResponse(
                    general_message='Fix the errors and try again ',
                    response={"Success": [], "Failed": error_rows}
                    ).get_failure_response()
            if len(voucher_serializer.data) != len(valid_rows):
                transaction.set_rollback(True)
                return CustomResponse(
                    general_message='Something went wrong. Please try again.').get_failure_response()

        for voucher in voucher_serializer.data:
            muid = voucher['muid']
            code = voucher['code']
            month = voucher['month']
            week = voucher['week']
            karma = voucher['karma']
            task_hashtag = voucher['hashtag']
            full_name = voucher['fullname']
            email = voucher['email']
            description = voucher['description']
            time_or_event = f'{month}/{week}'
            event = voucher['event']
            if event != '':
                time_or_event = f'{event}/{description}'

            success_rows.append({
                'muid': muid,
                'code': code,
                'user': full_name,
                'task': task_hashtag,
                'karma': karma,
                'month': month,
                'week': week,
                'description': description,
                'event': event
            })
            # Preparing email context and attachment
            from_mail = decouple.config("FROM_MAIL")
            subject = "Congratulations on earning Karma points!"
            text = f"""Greetings from GTech µLearn!

            Great news! You are just one step away from claiming your internship/contribution Karma points.

            Name: {full_name}
            Email: {email}

            To claim your karma points copy this `voucher {code}` and paste it #task-dropbox channel along with your voucher image.
            """

            karma_voucher_image = generate_karma_voucher(
                name=str(full_name), karma=str(int(karma)), code=code, hashtag=task_hashtag,
                month=time_or_event)
            karma_voucher_image.seek(0)
            email_obj = EmailMessage(
                subject=subject,
                body=text,
                from_email=from_mail,
                to=[email],
            )
            attachment = MIMEImage(karma_voucher_image.read())
            attachment.add_header(
                'Content-Disposition',
                'attachment',
                filename=f'{str(full_name)}.jpg',
            )
            email_obj.attach(attachment)
            email_obj.send(fail_silently=False)

        return CustomResponse(
            response={"Success": success_rows, "Failed": error_rows}
        ).get_success_response()


class VoucherLogAPI(APIView):
    authentication_classes = [CustomizePermission]

    @role_required([RoleType.ADMIN.value, RoleType.FELLOW.value, RoleType.ASSOCIATE.value])
    def get(self, request):
        voucher_queryset = VoucherLog.objects.all()
        paginated_queryset = CommonUtils.get_paginated_queryset(
            voucher_queryset, request,
            search_fields=["user__full_name",
                           "task__title", "karma", "month", "week", "claimed",
                           "updated_by__full_name",
                           "created_by__full_name",
                           "description", "event", "code"],

            sort_fields={'user': 'user__full_name',
                         'code': 'code',
                         'karma': 'karma',
                         'claimed': 'claimed',
                         'task': 'task__title',
                         'week': 'week',
                         'month': 'month',
                         'updated_by': 'updated_by__full_name',
                         'updated_at': 'updated_at',
                         'created_at': 'created_at',
                         'event': 'event',
                         'description': 'description'
                         }
        )
        voucher_serializer = VoucherLogSerializer(paginated_queryset.get('queryset'), many=True).data
        return CustomResponse().paginated_response(data=voucher_serializer,
                                                   pagination=paginated_queryset.get('pagination'))

    @role_required([RoleType.ADMIN.value, RoleType.FELLOW.value, RoleType.ASSOCIATE.value])
    def post(self, request):
        serializer = VoucherLogCreateSerializer(data=request.data, context={'request': request})
        if serializer.is_valid():
            with transaction.atomic():
                id = serializer.save().id
                voucher = VoucherLog.objects.filter(id=id).values(
                    'code',
                    'user__full_name',
                    'user__email',
                    'task__hashtag',
                    'month',
                    'week',
                    'karma'
                ).first()
                if not voucher:
                    transaction.set_rollback(True)
                    return CustomResponse(
                        general_message='Something went wrong. Please try again.').get_failure_response()
            code = voucher['code']
            month = voucher['month']
            week = voucher['week']
            karma = voucher['karma']
            task_hashtag = voucher['task__hashtag']
            full_name = voucher['user__full_name']
            email = voucher['user__email']

            # Preparing email context and attachment
            from_mail = decouple.config("FROM_MAIL")
            subject = "Congratulations on earning Karma points!"
            text = f"""Greetings from GTech µLearn!

            Great news! You are just one step away from claiming your internship/contribution Karma points.
            
            Name: {full_name}
            Email: {email}
            
            To claim your karma points copy this `voucher {code}` and paste it #task-dropbox channel along with your voucher image.
            """

            month_week = f'{month}/{week}'
            karma_voucher_image = generate_karma_voucher(
                name=str(full_name), karma=str(int(karma)), code=code, hashtag=task_hashtag,
                month=month_week)
            karma_voucher_image.seek(0)
            email_obj = EmailMessage(
                subject=subject,
                body=text,
                from_email=from_mail,
                to=[email],
            )
            attachment = MIMEImage(karma_voucher_image.read())
            attachment.add_header(
                'Content-Disposition',
                'attachment',
                filename=f'{str(full_name)}.jpg',
            )
            email_obj.attach(attachment)
            email_obj.send(fail_silently=False)
            return CustomResponse(general_message='Voucher created successfully',
                                  response=serializer.data).get_success_response()
        return CustomResponse(message=serializer.errors).get_failure_response()

    @role_required([RoleType.ADMIN.value, RoleType.FELLOW.value, RoleType.ASSOCIATE.value])
    def patch(self, request, voucher_id):
        user_id = JWTUtils.fetch_user_id(request)
        context = {'user_id': user_id}
        voucher = VoucherLog.objects.filter(id=voucher_id, claimed=False).first()
        if not voucher:
            return CustomResponse(general_message="Voucher Not available").get_failure_response()
        serializer = VoucherLogUpdateSerializer(voucher, data=request.data, context=context)
        if serializer.is_valid():
            serializer.save()
            return CustomResponse(general_message='Voucher updated successfully').get_success_response()
        return CustomResponse(message=serializer.errors).get_failure_response()

    @role_required([RoleType.ADMIN.value, RoleType.FELLOW.value, RoleType.ASSOCIATE.value])
    def delete(self, request, voucher_id):
        if voucher_log := VoucherLog.objects.filter(id=voucher_id).first():
            voucher_log.delete()
            return CustomResponse(
                general_message=f'Voucher successfully deleted'
            ).get_success_response()
        return CustomResponse(
            general_message=f'Invalid Voucher'
        ).get_failure_response()


class ExportVoucherLogAPI(APIView):
    authentication_classes = [CustomizePermission]

    @role_required([RoleType.ADMIN.value, RoleType.FELLOW.value, RoleType.ASSOCIATE.value])
    def get(self, request):
        voucher_serializer = VoucherLog.objects.all()
        voucher_serializer_data = VoucherLogSerializer(voucher_serializer, many=True).data

        return CommonUtils.generate_csv(voucher_serializer_data, 'Voucher Log')


class VoucherBaseTemplateAPI(APIView):
    authentication_classes = [CustomizePermission]

    def get(self, request):
        wb = load_workbook('./api/dashboard/karma_voucher/assets/voucher_base_template.xlsx')
        ws = wb['Data Definitions']
        hashtags = TaskList.objects.all().values_list('hashtag', flat=True)
        data = {
            'hashtag': hashtags,
            'month': ['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September',
                      'October', 'November', 'December'],
            'week': ['W1', 'W2', 'W3', 'W4', 'W5']
        }
        # Write data column-wise
        for col_num, (col_name, col_values) in enumerate(data.items(), start=1):
            for row, value in enumerate(col_values, start=2):
                ws.cell(row=row, column=col_num, value=value)
        # Save the file
        with NamedTemporaryFile() as tmp:
            tmp.close()  # with statement opened tmp, close it so wb.save can open it
            wb.save(tmp.name)
            with open(tmp.name, 'rb') as f:
                f.seek(0)
                new_file_object = f.read()
        return FileResponse(BytesIO(new_file_object), as_attachment=True, filename='voucher_base_template.xlsx')
