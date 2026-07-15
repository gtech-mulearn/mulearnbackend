import json
import os
import uuid
from io import BytesIO

import openpyxl
from django.conf import settings
from django.core.exceptions import ValidationError
from django.db.models import Q
from django.http import FileResponse
from django.utils.timezone import now
from rest_framework.generics import get_object_or_404
from rest_framework.parsers import FormParser, MultiPartParser, JSONParser
from rest_framework.views import APIView

from datetime import datetime, timedelta
from api.dashboard.achievement import achievement_serializer
from db.achievement import Achievement, AchievementRule, UserAchievementsLog, AchievementAuditLog
from db.task import Level
from db.user import User
from mu_celery.achievement_tasks import bulk_check_and_issue_achievements
from utils.permission import CustomizePermission, JWTUtils, RoleRequired, BackendApiKeyPermission
from utils.response import CustomResponse
from utils.types import RoleType
from utils.utils import DateTimeUtils, CommonUtils
from drf_spectacular.utils import extend_schema, inline_serializer, OpenApiResponse
from rest_framework import serializers as s


class AchievementListAPIView(APIView):
    @extend_schema(
        tags=['Dashboard - Achievement'],
        description="Retrieve Achievement List. Pass ?user_id=<muid> to include a 'has_achievement' flag per item.",
        responses={200: achievement_serializer.AchievementSerializer},
    )
    def get(self, request):
        user_id = JWTUtils.fetch_user_id(request)

        if not user_id:
            return CustomResponse(
                general_message="Invalid or missing token"
            ).get_failure_response()

        user = User.objects.filter(id=user_id).first()

        if not user:
            return CustomResponse(
                general_message="User Not Exists"
            ).get_failure_response()

        # Optional: check which achievements a specific user already has
        target_user_id = request.query_params.get('user_id')
        user_achievement_ids = set()
        if target_user_id:
            user_achievement_ids = set(
                UserAchievementsLog.objects.filter(
                    user_id=target_user_id
                ).values_list('achievement_id_id', flat=True)
            )

        achievements = Achievement.objects.all()
        achievements_serializer = achievement_serializer.AchievementSerializer(
            achievements,
            many=True,
            context={
                'request': request,
                'user_achievements': user_achievement_ids,
            }
        )

        return CustomResponse(
            response=achievements_serializer.data
        ).get_success_response()


class AchievementCreateAPIView(APIView):
    parser_classes = [MultiPartParser, FormParser, JSONParser]
    
    @extend_schema(tags=['Dashboard - Achievement'], description="Create Achievement Create.",
        responses={200: achievement_serializer.AchievementSerializer},
    )
    def post(self, request):
        user_id = JWTUtils.fetch_user_id(request)

        if not user_id:
            return CustomResponse(
                general_message="Invalid or missing token"
            ).get_failure_response()

        user = User.objects.filter(id=user_id).first()
        if not user:
            return CustomResponse(
                general_message="User Not Exists"
            ).get_failure_response()

        data = request.data
        # Icon can be either a file upload or a text URL
        icon_file = request.FILES.get("icon")
        icon_url = data.get("icon", "") if not icon_file else ""
        
        required_fields = ["name", "description", "tags", "type", "has_vc"]
        missing_fields = [field for field in required_fields if field not in data]
        if missing_fields:
            return CustomResponse(
                general_message=f"Missing required fields: {', '.join(missing_fields)}"
            ).get_failure_response()

        # Parse has_vc from string to boolean (FormData sends strings)
        has_vc_value = data.get("has_vc")
        if isinstance(has_vc_value, str):
            has_vc_value = has_vc_value.lower() in ("true", "1", "yes")
        
        tags_value = data.get("tags", [])
        if isinstance(tags_value, str):
            try:
                tags_value = json.loads(tags_value)
            except json.JSONDecodeError:
                tags_value = []

        if Achievement.objects.filter(name=data["name"]).exists():
            return CustomResponse(
                general_message="Name already exists"
            ).get_failure_response()

        level = None
        if "level_id" in data and data["level_id"]:
            try:
                level = Level.objects.get(id=data["level_id"])
            except Level.DoesNotExist:
                return CustomResponse(
                    general_message="Invalid level_id"
                ).get_failure_response()

        # Handle icon file upload
        icon_path = icon_url  # Default to URL if provided
        if icon_file:
            # Validate file type
            allowed_extensions = ['jpg', 'jpeg', 'png', 'gif', 'webp', 'svg']
            file_ext = icon_file.name.split('.')[-1].lower()
            if file_ext not in allowed_extensions:
                return CustomResponse(
                    general_message=f"Invalid file type. Allowed: {', '.join(allowed_extensions)}"
                ).get_failure_response()
            
            # Validate file size (max 5MB)
            if icon_file.size > 5 * 1024 * 1024:
                return CustomResponse(
                    general_message="File size exceeds 5MB limit"
                ).get_failure_response()
            
            # Create directory if it doesn't exist
            upload_dir = os.path.join(settings.MEDIA_ROOT, 'achievements', 'icons')
            os.makedirs(upload_dir, exist_ok=True)
            
            # Generate unique filename
            unique_filename = f"{uuid.uuid4()}.{file_ext}"
            file_path = os.path.join(upload_dir, unique_filename)
            
            # Save file
            with open(file_path, 'wb+') as destination:
                for chunk in icon_file.chunks():
                    destination.write(chunk)
            
            # Store relative path for database
            icon_path = f"achievements/icons/{unique_filename}"

        achievement = Achievement.objects.create(
            id=str(uuid.uuid4()),
            name=data["name"],
            description=data["description"],
            icon=icon_path,
            tags=tags_value,
            type=data["type"],
            level_id=level,
            has_vc=has_vc_value,
            template_id=data.get("template_id"),
            created_by=user,
            updated_by=user,
            created_at=now(),
            updated_at=now(),
        )

        return CustomResponse(
            general_message=f"Achievement '{achievement.name}' created successfully!"
        ).get_success_response()


class AchievementUpdateAPIView(APIView):
    parser_classes = [MultiPartParser, FormParser, JSONParser]

    @extend_schema(
        tags=['Dashboard - Achievement'],
        description="Update Achievement Update.",
        responses={200: achievement_serializer.AchievementSerializer},
    )
    def put(self, request, achievement_id=None):
        user_id = JWTUtils.fetch_user_id(request)

        if not user_id:
            return CustomResponse(
                general_message="Invalid or missing token"
            ).get_failure_response()

        user = User.objects.filter(id=user_id).first()
        if not user:
            return CustomResponse(
                general_message="User Not Exists"
            ).get_failure_response()

        if not achievement_id:
            return CustomResponse(
                general_message="Achievement ID is required"
            ).get_failure_response()

        try:
            achievement = Achievement.objects.get(id=achievement_id)
        except Achievement.DoesNotExist:
            return CustomResponse(
                general_message="Achievement not found"
            ).get_failure_response()

        # Convert QueryDict to regular dict to properly handle list/boolean assignments
        data = dict(request.data)
        # QueryDict wraps values in lists, so unwrap single values
        # Exclude intentional list fields like 'tags' from unwrapping
        list_fields = {'tags'}
        for key in data:
            if key not in list_fields and isinstance(data[key], list) and len(data[key]) == 1:
                data[key] = data[key][0]
        data["updated_by"] = user_id

        # Handle icon file upload
        icon_file = request.FILES.get("icon")
        if icon_file:
            # Validate file type
            allowed_extensions = ['jpg', 'jpeg', 'png', 'gif', 'webp', 'svg']
            file_ext = icon_file.name.split('.')[-1].lower()
            if file_ext not in allowed_extensions:
                return CustomResponse(
                    general_message=f"Invalid file type. Allowed: {', '.join(allowed_extensions)}"
                ).get_failure_response()
            
            # Validate file size (max 5MB)
            if icon_file.size > 5 * 1024 * 1024:
                return CustomResponse(
                    general_message="File size exceeds 5MB limit"
                ).get_failure_response()
            
            # Create directory if it doesn't exist
            upload_dir = os.path.join(settings.MEDIA_ROOT, 'achievements', 'icons')
            os.makedirs(upload_dir, exist_ok=True)
            
            # Generate unique filename
            unique_filename = f"{uuid.uuid4()}.{file_ext}"
            file_path = os.path.join(upload_dir, unique_filename)
            
            # Save file
            with open(file_path, 'wb+') as destination:
                for chunk in icon_file.chunks():
                    destination.write(chunk)
            
            # Store relative path for database
            data["icon"] = f"achievements/icons/{unique_filename}"
        elif "icon" not in data or not data["icon"]:
            # Keep existing icon if no new file or URL provided
            data["icon"] = achievement.icon

        if "level_id" in data:
            if data["level_id"]:
                try:
                    level = Level.objects.get(id=data["level_id"])
                    data["level_id"] = level.id
                except Level.DoesNotExist:
                    return CustomResponse(
                        general_message="Invalid level_id"
                    ).get_failure_response()
            else:
                data["level_id"] = None

        # Parse has_vc from string to boolean (FormData sends strings)
        if "has_vc" in data:
            has_vc_value = data.get("has_vc")
            if isinstance(has_vc_value, str):
                data["has_vc"] = has_vc_value.lower() in ("true", "1", "yes")
        
        # Parse tags from JSON string to list (FormData sends strings)
        if "tags" in data:
            tags_value = data.get("tags", [])
            if isinstance(tags_value, str):
                try:
                    data["tags"] = json.loads(tags_value)
                except json.JSONDecodeError:
                    data["tags"] = []

        serializer = achievement_serializer.AchievementSerializer(
            achievement, data=data, partial=True
        )

        if serializer.is_valid():
            serializer.save()
            return CustomResponse(
                general_message="Achievement updated successfully"
            ).get_success_response()

        return CustomResponse(
            general_message="Invalid Data", response=serializer.errors
        ).get_failure_response()


class AchievementDeleteAPIView(APIView):
    @extend_schema(tags=['Dashboard - Achievement'], description="Delete Achievement Delete.",
        responses={200: achievement_serializer.AchievementSerializer},
    )
    def delete(self, request, achievement_id):
        user_id = JWTUtils.fetch_user_id(request)

        if not user_id:
            return CustomResponse(
                general_message="Invalid or missing token"
            ).get_failure_response()

        if not User.objects.filter(id=user_id).exists():
            return CustomResponse(
                general_message="User Not Exists"
            ).get_failure_response()

        try:
            achievement = Achievement.objects.get(id=achievement_id)
        except Achievement.DoesNotExist:
            return CustomResponse(
                general_message="Achievement not found"
            ).get_failure_response()

        achievement.delete()
        return CustomResponse(
            general_message="Achievement deleted successfully"
        ).get_success_response()


class UserAchievementsListAPIView(APIView):
    @extend_schema(
        tags=['Dashboard - Achievement'],
        description="Retrieve User Achievements List.",
        responses={200: achievement_serializer.UserAchievementsSerializer},
    )
    def get(self, request, muid):
        try:
            user = get_object_or_404(User, muid=muid)

            user_achievements = (
                UserAchievementsLog.objects.filter(user_id=user.id)
                .select_related("achievement_id")
                .only(
                    "id",
                    "user_id",
                    "achievement_id",
                    "is_issued",
                    "vc_url",
                    "achievement_id__name",
                    "achievement_id__description",
                )
            )

            # if not user_achievements.exists():
            #     return CustomResponse(
            #         general_message="No achievements found for this user"
            #     ).get_failure_response()

            serializer = achievement_serializer.UserAchievementsSerializer(
                user_achievements, many=True
            )
            return CustomResponse(response=serializer.data).get_success_response()

        except ValidationError:
            return CustomResponse(
                general_message="Invalid format for muid"
            ).get_failure_response()

        except Exception as e:
            return CustomResponse(
                general_message=f"An unexpected error occurred: {str(e)}"
            ).get_failure_response()


class UserAchievementsIssueAPIView(APIView):
    @extend_schema(tags=['Dashboard - Achievement'], description="Create User Achievements Issue.",
        responses={200: achievement_serializer.AchievementSerializer},
    )
    def post(self, request):
        user_id = JWTUtils.fetch_user_id(request)
        if not user_id:
            return CustomResponse(
                general_message="Invalid or missing token"
            ).get_failure_response()

        achievement_id = request.data.get("achievement_id")
        vc_url = request.data.get("vc_url")

        if not achievement_id:
            return CustomResponse(
                general_message="Achievement ID is required"
            ).get_failure_response()

        if not vc_url:
            return CustomResponse(
                general_message="VC URL is required"
            ).get_failure_response()

        if not User.objects.filter(id=user_id).exists():
            return CustomResponse(
                general_message="User Not Exists"
            ).get_failure_response()

        try:
            user_achievement = UserAchievementsLog.objects.get(
                user_id=user_id, achievement_id=achievement_id
            )
        except UserAchievementsLog.DoesNotExist:
            return CustomResponse(
                general_message="Achievement record not found"
            ).get_failure_response()

        if user_achievement.vc_url:
            return CustomResponse(
                general_message="This achievement has already been issued"
            ).get_failure_response()

        UserAchievementsLog.objects.filter(
            user_id=user_id, achievement_id=achievement_id
        ).update(is_issued=True, vc_url=vc_url)

        return CustomResponse(
            general_message="Achievement issued successfully"
        ).get_success_response()


# ============================================================================
# NEW ACHIEVEMENT SYSTEM VIEWS
# ============================================================================


class EligibleAchievementsAPIView(APIView):
    """Get achievements the current user is eligible to claim"""

    @extend_schema(tags=['Dashboard - Achievement'], description="Retrieve Eligible Achievements.",
        responses={200: achievement_serializer.AchievementSerializer},
    )
    def get(self, request):
        user_id = JWTUtils.fetch_user_id(request)
        if not user_id:
            return CustomResponse(
                general_message="Invalid or missing token"
            ).get_failure_response()

        from api.dashboard.achievement.rule_engine import RuleEvaluator

        evaluator = RuleEvaluator(user_id)
        eligible = evaluator.get_eligible_achievements()

        response_data = [
            {
                "achievement_id": result.achievement_id,
                "achievement_name": result.achievement_name,
                "eligible": result.eligible,
                "reason": result.reason,
                "progress": result.progress,
            }
            for result in eligible
        ]

        return CustomResponse(response=response_data).get_success_response()


class ClaimAchievementAPIView(APIView):
    """Claim an achievement (user action)"""

    @extend_schema(tags=['Dashboard - Achievement'], description="Create Claim Achievement.",
        responses={200: achievement_serializer.AchievementSerializer},
    )
    def post(self, request, achievement_id):
        user_id = JWTUtils.fetch_user_id(request)
        if not user_id:
            return CustomResponse(
                general_message="Invalid or missing token"
            ).get_failure_response()

        from mu_celery.achievement_tasks import claim_achievement

        result = claim_achievement(user_id, achievement_id)

        if result["success"]:
            return CustomResponse(
                general_message=result["message"],
                response={
                    "achievement_name": result.get("achievement_name"),
                    "vc_pending": result.get("vc_pending", False),
                },
            ).get_success_response()
        else:
            return CustomResponse(
                general_message=result["message"],
                response={"progress": result.get("progress")},
            ).get_failure_response()


class UserProgressAPIView(APIView):
    """Get progress towards all achievements"""

    @extend_schema(
        tags=['Dashboard - Achievement'], description="Retrieve User Progress.",
        responses={200: inline_serializer("AchievementUserProgressResponse", fields={
            "hasError": s.BooleanField(default=False),
            "statusCode": s.IntegerField(default=200),
            "message": s.DictField(default={}),
            "response": s.ListField(
                child=inline_serializer("AchievementProgressItem", fields={
                    "achievement_id": s.CharField(),
                    "achievement_name": s.CharField(),
                    "eligible": s.BooleanField(),
                    "claimed": s.BooleanField(),
                    "reason": s.CharField(allow_null=True),
                    "progress": s.DictField(
                        help_text="Rule-engine progress data (e.g. current vs. required counts)"
                    ),
                }),
                help_text="Progress towards every achievement for the current user",
            ),
        })},
    )
    def get(self, request):
        user_id = JWTUtils.fetch_user_id(request)
        if not user_id:
            return CustomResponse(
                general_message="Invalid or missing token"
            ).get_failure_response()

        from api.dashboard.achievement.rule_engine import RuleEvaluator

        evaluator = RuleEvaluator(user_id)
        all_progress = evaluator.get_all_progress()

        response_data = [
            {
                "achievement_id": result.achievement_id,
                "achievement_name": result.achievement_name,
                "eligible": result.eligible,
                "claimed": result.claimed,
                "reason": result.reason,
                "progress": result.progress,
            }
            for result in all_progress
        ]

        return CustomResponse(response=response_data).get_success_response()


class AchievementRuleListAPIView(APIView):
    """List all achievement rules (admin)"""

    @extend_schema(tags=['Dashboard - Achievement'], description="Retrieve Achievement Rule List.",
        responses={200: achievement_serializer.AchievementSerializer},
    )
    def get(self, request):
        user_id = JWTUtils.fetch_user_id(request)
        if not user_id:
            return CustomResponse(
                general_message="Invalid or missing token"
            ).get_failure_response()

        rules = AchievementRule.objects.all().select_related("achievement")
        data = [
            {
                "id": str(rule.id),
                "achievement_id": str(rule.achievement_id),
                "achievement_name": rule.achievement.name,
                "version": rule.version,
                "rule_type": rule.rule_type,
                "conditions": rule.conditions,
                "is_active": rule.is_active,
                "created_at": rule.created_at.isoformat() if rule.created_at else None,
            }
            for rule in rules
        ]

        return CustomResponse(response=data).get_success_response()


class AchievementRuleCreateAPIView(APIView):
    """Create a new achievement rule (admin)"""

    @extend_schema(tags=['Dashboard - Achievement'], description="Create Achievement Rule Create.",
        responses={200: achievement_serializer.AchievementSerializer},
    )
    def post(self, request):
        user_id = JWTUtils.fetch_user_id(request)
        if not user_id:
            return CustomResponse(
                general_message="Invalid or missing token"
            ).get_failure_response()

        data = request.data
        required_fields = ["achievement_id", "rule_type", "conditions"]
        missing_fields = [field for field in required_fields if field not in data]
        if missing_fields:
            return CustomResponse(
                general_message=f"Missing required fields: {', '.join(missing_fields)}"
            ).get_failure_response()

        # Get next version number
        existing_versions = AchievementRule.objects.filter(
            achievement_id=data["achievement_id"]
        ).values_list("version", flat=True)
        next_version = max(existing_versions) + 1 if existing_versions else 1

        # Deactivate existing rules for this achievement
        AchievementRule.objects.filter(
            achievement_id=data["achievement_id"], is_active=True
        ).update(is_active=False)

        rule = AchievementRule.objects.create(
            id=str(uuid.uuid4()),
            achievement_id=data["achievement_id"],
            version=next_version,
            rule_type=data["rule_type"],
            conditions=data["conditions"],
            is_active=True,
            created_by_id=user_id,
            created_at=now(),
            updated_at=now(),
        )

        return CustomResponse(
            general_message=f"Rule v{next_version} created successfully",
            response={"rule_id": str(rule.id), "version": next_version},
        ).get_success_response()


class AchievementRuleDetailAPIView(APIView):
    """Get or update details of a specific rule (admin)"""

    @extend_schema(tags=['Dashboard - Achievement'], description="Retrieve Achievement Rule Detail.",
        responses={200: achievement_serializer.AchievementSerializer},
    )
    def get(self, request, rule_id):
        user_id = JWTUtils.fetch_user_id(request)
        if not user_id:
            return CustomResponse(
                general_message="Invalid or missing token"
            ).get_failure_response()

        try:
            rule = AchievementRule.objects.select_related("achievement").get(id=rule_id)
        except AchievementRule.DoesNotExist:
            return CustomResponse(
                general_message="Rule not found"
            ).get_failure_response()

        data = {
            "id": str(rule.id),
            "achievement_id": str(rule.achievement_id),
            "achievement_name": rule.achievement.name,
            "version": rule.version,
            "rule_type": rule.rule_type,
            "conditions": rule.conditions,
            "is_active": rule.is_active,
            "created_at": rule.created_at.isoformat() if rule.created_at else None,
        }

        return CustomResponse(response=data).get_success_response()

    @extend_schema(
        tags=['Dashboard - Achievement'],
        description="Update a rule's rule_type and/or conditions. Works for both active and deactivated rules. Version and achievement association are immutable.",
        responses={200: achievement_serializer.AchievementSerializer},
    )
    def patch(self, request, rule_id):
        user_id = JWTUtils.fetch_user_id(request)
        if not user_id:
            return CustomResponse(
                general_message="Invalid or missing token"
            ).get_failure_response()

        try:
            rule = AchievementRule.objects.select_related("achievement").get(id=rule_id)
        except AchievementRule.DoesNotExist:
            return CustomResponse(
                general_message="Rule not found"
            ).get_failure_response()

        EDITABLE_FIELDS = {"rule_type", "conditions"}
        data = request.data

        unknown_fields = set(data.keys()) - EDITABLE_FIELDS
        if unknown_fields:
            return CustomResponse(
                general_message=f"Fields not editable: {', '.join(sorted(unknown_fields))}. Only rule_type and conditions can be updated."
            ).get_failure_response()

        if not EDITABLE_FIELDS.intersection(data.keys()):
            return CustomResponse(
                general_message="No editable fields provided. Supply at least one of: rule_type, conditions."
            ).get_failure_response()

        if "rule_type" in data:
            valid_rule_types = [choice[0] for choice in AchievementRule.RULE_TYPE_CHOICES]
            if data["rule_type"] not in valid_rule_types:
                return CustomResponse(
                    general_message=f"Invalid rule_type '{data['rule_type']}'. Valid choices: {', '.join(valid_rule_types)}"
                ).get_failure_response()
            rule.rule_type = data["rule_type"]

        if "conditions" in data:
            if not isinstance(data["conditions"], dict):
                return CustomResponse(
                    general_message="conditions must be a JSON object."
                ).get_failure_response()
            rule.conditions = data["conditions"]

        rule.save(update_fields=[f for f in EDITABLE_FIELDS if f in data] + ["updated_at"])

        return CustomResponse(
            general_message=f"Rule v{rule.version} updated successfully",
            response={
                "id": str(rule.id),
                "achievement_id": str(rule.achievement_id),
                "achievement_name": rule.achievement.name,
                "version": rule.version,
                "rule_type": rule.rule_type,
                "conditions": rule.conditions,
                "is_active": rule.is_active,
            },
        ).get_success_response()


class AchievementRuleDeactivateAPIView(APIView):
    """Deactivate a rule (admin)"""

    @extend_schema(tags=['Dashboard - Achievement'], description="Create Achievement Rule Deactivate.",
        responses={200: achievement_serializer.AchievementSerializer},
    )
    def post(self, request, rule_id):
        user_id = JWTUtils.fetch_user_id(request)
        if not user_id:
            return CustomResponse(
                general_message="Invalid or missing token"
            ).get_failure_response()

        try:
            rule = AchievementRule.objects.get(id=rule_id)
        except AchievementRule.DoesNotExist:
            return CustomResponse(
                general_message="Rule not found"
            ).get_failure_response()

        rule.is_active = False
        rule.save()

        return CustomResponse(
            general_message=f"Rule v{rule.version} deactivated"
        ).get_success_response()


class AchievementRuleActivateAPIView(APIView):
    """Activate a rule (admin)"""

    @extend_schema(tags=['Dashboard - Achievement'], description="Activate an Achievement Rule.",
        responses={200: achievement_serializer.AchievementSerializer},
    )
    def post(self, request, rule_id):
        user_id = JWTUtils.fetch_user_id(request)
        if not user_id:
            return CustomResponse(
                general_message="Invalid or missing token"
            ).get_failure_response()

        try:
            rule = AchievementRule.objects.get(id=rule_id)
        except AchievementRule.DoesNotExist:
            return CustomResponse(
                general_message="Rule not found"
            ).get_failure_response()

        if rule.is_active:
            return CustomResponse(
                general_message=f"Rule v{rule.version} is already active"
            ).get_failure_response()

        rule.is_active = True
        rule.save()

        return CustomResponse(
            general_message=f"Rule v{rule.version} activated"
        ).get_success_response()


class SimulateRulesAPIView(APIView):
    """Simulate rule evaluation for a user (admin/debug)"""

    @extend_schema(tags=['Dashboard - Achievement'], description="Retrieve Simulate Rules.",
        responses={200: inline_serializer("AchievementSimulateRulesResponse", fields={
            "hasError": s.BooleanField(default=False),
            "statusCode": s.IntegerField(default=200),
            "message": s.DictField(default={}),
            "response": s.ListField(
                child=inline_serializer("AchievementSimulateItem", fields={
                    "achievement_id": s.CharField(),
                    "achievement_name": s.CharField(),
                    "eligible": s.BooleanField(),
                    "claimed": s.BooleanField(),
                    "reason": s.CharField(allow_null=True),
                    "progress": s.DictField(
                        help_text="Rule-engine progress data for the target user"
                    ),
                }),
                help_text="Simulated rule evaluation results for the given user (by muid)",
            ),
        })},
    )
    def get(self, request, muid):
        user_id = JWTUtils.fetch_user_id(request)
        if not user_id:
            return CustomResponse(
                general_message="Invalid or missing token"
            ).get_failure_response()

        try:
            target_user = User.objects.get(muid=muid)
        except User.DoesNotExist:
            return CustomResponse(
                general_message="User not found"
            ).get_failure_response()

        from api.dashboard.achievement.rule_engine import RuleEvaluator

        evaluator = RuleEvaluator(str(target_user.id))
        all_progress = evaluator.get_all_progress()

        response_data = [
            {
                "achievement_id": result.achievement_id,
                "achievement_name": result.achievement_name,
                "eligible": result.eligible,
                "claimed": result.claimed,
                "reason": result.reason,
                "progress": result.progress,
            }
            for result in all_progress
        ]

        return CustomResponse(response=response_data).get_success_response()


class DebugAchievementAPIView(APIView):
    """Debug a specific achievement for a user (admin)"""

    @extend_schema(tags=['Dashboard - Achievement'], description="Retrieve Debug Achievement.",
        responses={200: achievement_serializer.AchievementSerializer},
    )
    def get(self, request, muid, achievement_id):
        user_id = JWTUtils.fetch_user_id(request)
        if not user_id:
            return CustomResponse(
                general_message="Invalid or missing token"
            ).get_failure_response()

        try:
            target_user = User.objects.get(muid=muid)
        except User.DoesNotExist:
            return CustomResponse(
                general_message="User not found"
            ).get_failure_response()

        from api.dashboard.achievement.rule_engine import RuleEvaluator
        from db.achievement import UserIgKarma, UserStreak, UserSkillProgress

        evaluator = RuleEvaluator(str(target_user.id))
        result = evaluator.evaluate_achievement(achievement_id)

        if not result:
            return CustomResponse(
                general_message="No active rule found for this achievement"
            ).get_failure_response()

        # Get additional debug data
        ig_karma = list(
            UserIgKarma.objects.filter(user_id=target_user.id).values(
                "ig_id", "total_karma", "task_count"
            )
        )
        streaks = list(
            UserStreak.objects.filter(user_id=target_user.id).values(
                "streak_type", "current_streak", "longest_streak"
            )
        )
        skill_progress = list(
            UserSkillProgress.objects.filter(user_id=target_user.id).values(
                "skill_id", "completed_task_count", "total_karma"
            )
        )

        response_data = {
            "evaluation": {
                "achievement_id": result.achievement_id,
                "achievement_name": result.achievement_name,
                "eligible": result.eligible,
                "reason": result.reason,
                "progress": result.progress,
            },
            "user_data": {
                "ig_karma": ig_karma,
                "streaks": streaks,
                "skill_progress": skill_progress,
            },
        }

        return CustomResponse(response=response_data).get_success_response()


class ManualIssueAPIView(APIView):
    """Manually issue an achievement (admin)"""

    @extend_schema(tags=['Dashboard - Achievement'], description="Create Manual Issue.",
        responses={200: OpenApiResponse(description="Achievement manually issued to user")},
    )
    def post(self, request):
        user_id = JWTUtils.fetch_user_id(request)
        if not user_id:
            return CustomResponse(
                general_message="Invalid or missing token"
            ).get_failure_response()

        muid = request.data.get("muid")
        achievement_id = request.data.get("achievement_id")

        if not muid or not achievement_id:
            return CustomResponse(
                general_message="muid and achievement_id are required"
            ).get_failure_response()

        try:
            target_user = User.objects.get(muid=muid)
        except User.DoesNotExist:
            return CustomResponse(
                general_message="User not found"
            ).get_failure_response()

        from mu_celery.achievement_tasks import manual_issue_achievement

        result = manual_issue_achievement(
            user_id=str(target_user.id),
            achievement_id=achievement_id,
            performed_by=user_id,
        )

        if result["success"]:
            return CustomResponse(
                general_message=result["message"]
            ).get_success_response()
        else:
            return CustomResponse(
                general_message=result["message"]
            ).get_failure_response()


class RevokeAchievementAPIView(APIView):
    """Revoke an achievement (admin)"""

    @extend_schema(tags=['Dashboard - Achievement'], description="Create Revoke Achievement.",
        responses={200: achievement_serializer.AchievementSerializer},
    )
    def post(self, request):
        user_id = JWTUtils.fetch_user_id(request)
        if not user_id:
            return CustomResponse(
                general_message="Invalid or missing token"
            ).get_failure_response()

        muid = request.data.get("muid")
        achievement_id = request.data.get("achievement_id")
        reason = request.data.get("reason")

        if not muid or not achievement_id:
            return CustomResponse(
                general_message="muid and achievement_id are required"
            ).get_failure_response()

        try:
            target_user = User.objects.get(muid=muid)
        except User.DoesNotExist:
            return CustomResponse(
                general_message="User not found"
            ).get_failure_response()

        from mu_celery.achievement_tasks import revoke_achievement

        result = revoke_achievement(
            user_id=str(target_user.id),
            achievement_id=achievement_id,
            performed_by=user_id,
            reason=reason,
        )

        if result["success"]:
            return CustomResponse(
                general_message=result["message"]
            ).get_success_response()
        else:
            return CustomResponse(
                general_message=result["message"]
            ).get_failure_response()


class AuditLogAPIView(APIView):
    """View audit logs for a user (admin)"""

    @extend_schema(tags=['Dashboard - Achievement'], description="Retrieve Audit Log.",
        responses={200: inline_serializer("AchievementAuditLogResponse", fields={
            "hasError": s.BooleanField(default=False),
            "statusCode": s.IntegerField(default=200),
            "message": s.DictField(default={}),
            "response": s.ListField(
                child=inline_serializer("AchievementAuditLogItem", fields={
                    "id": s.CharField(),
                    "achievement_id": s.CharField(),
                    "achievement_name": s.CharField(),
                    "action": s.CharField(help_text="e.g. ISSUED, REVOKED"),
                    "rule_version": s.IntegerField(allow_null=True),
                    "metadata": s.DictField(allow_null=True),
                    "performed_by": s.CharField(allow_null=True),
                    "created_at": s.DateTimeField(allow_null=True),
                }),
                help_text="Last 100 audit log entries for the given user's achievements",
            ),
        })},
    )
    def get(self, request, muid):
        user_id = JWTUtils.fetch_user_id(request)
        if not user_id:
            return CustomResponse(
                general_message="Invalid or missing token"
            ).get_failure_response()

        try:
            target_user = User.objects.get(muid=muid)
        except User.DoesNotExist:
            return CustomResponse(
                general_message="User not found"
            ).get_failure_response()

        audit_logs = (
            AchievementAuditLog.objects.filter(user_id=target_user.id)
            .select_related("achievement")
            .order_by("-created_at")[:100]
        )

        data = [
            {
                "id": str(log.id),
                "achievement_id": str(log.achievement_id),
                "achievement_name": log.achievement.name,
                "action": log.action,
                "rule_version": log.rule_version,
                "metadata": log.metadata,
                "performed_by": str(log.performed_by_id) if log.performed_by_id else None,
                "created_at": log.created_at.isoformat() if log.created_at else None,
            }
            for log in audit_logs
        ]

        return CustomResponse(response=data).get_success_response()



class AchievementIssueBulkAPIView(APIView):
    from rest_framework.parsers import MultiPartParser, FormParser
    parser_classes = [MultiPartParser, FormParser]

    @extend_schema(tags=['Dashboard - Achievement'], description="Create Achievement Issue Bulk.",
        responses={200: achievement_serializer.AchievementSerializer},
    )
    def post(self, request):
        user_id = JWTUtils.fetch_user_id(request)
        if not user_id:
            return CustomResponse(
                general_message="Invalid or missing token"
            ).get_failure_response()

        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            return CustomResponse(
                general_message="No file uploaded"
            ).get_failure_response()

        try:
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active
            
            headers = [cell.value for cell in sheet[1]]
            required_headers = ['muid', 'achievement_id']
            
            if not all(h in headers for h in required_headers):
                return CustomResponse(
                    general_message=f"Missing required headers. Required: {required_headers}"
                ).get_failure_response()
            
            muid_idx = headers.index('muid')
            ach_idx = headers.index('achievement_id')
            
            success_count = 0
            failed_rows = []
            
            from mu_celery.achievement_tasks import manual_issue_achievement

            for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                muid = row[muid_idx]
                achievement_id = row[ach_idx]
                
                if not muid or not achievement_id:
                    continue
                     
                try:
                    user = User.objects.filter(muid=muid).first()
                    if not user:
                        failed_rows.append({"row": i, "muid": muid, "reason": "User not found"})
                        continue

                    result = manual_issue_achievement(
                        user_id=str(user.id),
                        achievement_id=str(achievement_id),
                        performed_by=user_id
                    )
                    
                    if result['success']:
                        success_count += 1
                    else:
                        failed_rows.append({"row": i, "muid": muid, "reason": result['message']})
                        
                except Exception as e:
                    failed_rows.append({"row": i, "muid": muid, "reason": str(e)})

            return CustomResponse(
                response={
                    "success_count": success_count,
                    "failed_count": len(failed_rows),
                    "failed_rows": failed_rows
                },
                general_message="Bulk issue processing completed"
            ).get_success_response()

        except Exception as e:
            return CustomResponse(
                general_message=f"Error processing file: {str(e)}"
            ).get_failure_response()


class AchievementBulkImportTemplateAPIView(APIView):
    @extend_schema(tags=['Dashboard - Achievement'], description="Retrieve Achievement Bulk Import Template.",
        responses={200: achievement_serializer.AchievementSerializer},
    )
    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['muid', 'achievement_id'])
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = FileResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=achievement_bulk_import_template.xlsx'
        return response


class AchievementLogListAPIView(APIView):
    @extend_schema(tags=['Dashboard - Achievement'], description="Retrieve Achievement Log List.",
        responses={200: achievement_serializer.AchievementSerializer},
    )
    def get(self, request):
        user_id = JWTUtils.fetch_user_id(request)
        if not user_id:
            return CustomResponse(
                general_message="Invalid or missing token"
            ).get_failure_response()
            
        logs = UserAchievementsLog.objects.select_related('user_id', 'achievement_id', 'updated_by').order_by('-created_at')
        
        paginated_queryset = CommonUtils.get_paginated_queryset(
            logs, 
            request, 
            search_fields=['user_id__muid', 'user_id__first_name', 'achievement_id__name'],
            sort_fields={'created_at': 'created_at'}
        )
        
        data = []
        for log in paginated_queryset.get('queryset'):
            data.append({
                "id": str(log.id),
                "muid": log.user_id.muid,
                "user_name": log.user_id.full_name,
                "achievement_name": log.achievement_id.name,
                "is_issued": log.is_issued,
                "created_at": log.created_at.isoformat() if log.created_at else None,
                "issued_by": log.updated_by.full_name if log.updated_by else None
            })
             
        return CustomResponse().paginated_response(
            data=data,
            pagination=paginated_queryset.get('pagination')
        )

class BulkClaimTaskAchievementAPIView(APIView):
    permission_classes = [BackendApiKeyPermission]

    @extend_schema(tags=['Dashboard - Achievement'], description="Create Bulk Claim Task Achievement.",
        responses={200: achievement_serializer.AchievementSerializer},
    )
    def post(self, request):
        try:
            today = datetime.now().date()
            yesterday = today - timedelta(days=1)

            date_from_str = request.data.get("date_from", yesterday.isoformat())
            date_to_str = request.data.get("date_to", yesterday.isoformat())

            # Validate date format and convert to date objects
            date_from = datetime.fromisoformat(date_from_str).date()
            date_to = datetime.fromisoformat(date_to_str).date()

            # Validate that the date range is logically correct
            if date_from > date_to:
                return CustomResponse(
                    general_message="Invalid date range: date_from cannot be after date_to."
                ).get_failure_response()
            bulk_check_and_issue_achievements.delay(
                date_from_str=date_from_str,
                date_to_str=date_to_str,
                performed_by_id=None,
            )

            return CustomResponse(
                general_message="Bulk sync job scheduled successfully."
            ).get_success_response()

        except ValueError:
            return CustomResponse(
                general_message="Invalid date format. Use ISO format (YYYY-MM-DD)."
            ).get_failure_response()

        except Exception as e:
            return CustomResponse(
                general_message=f"An unexpected error occurred: {str(e)}"
            ).get_failure_response()
