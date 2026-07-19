import uuid
from io import BytesIO

import openpyxl
from django.db import transaction
from django.db.models import Count
from django.http import FileResponse
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.views import APIView
from drf_spectacular.utils import OpenApiResponse, extend_schema

from db.intern import UserInternGuildLink
from db.user import User, UserRoleLink, Role
from utils.permission import CustomizePermission, JWTUtils, role_required
from utils.response import CustomResponse
from utils.types import RoleType, InternGuildStatus
from utils.utils import CommonUtils

from .serializers import ManageInternSerializer

class ManageInternAPI(APIView):
    authentication_classes = [CustomizePermission]

    @role_required([RoleType.ADMIN.value, RoleType.INTERN.value, RoleType.INTERN_LEAD.value])
    @extend_schema(
        tags=['Dashboard - Intern'],
        description="Retrieve intern(s) or intern detail.",
        responses={200: ManageInternSerializer(many=True)},
    )
    def get(self, request, intern_id=None):
        if intern_id:
            intern = UserInternGuildLink.objects.select_related('user').filter(id=intern_id).first()
            if not intern:
                return CustomResponse(general_message="Intern not found.").get_failure_response()
            serializer = ManageInternSerializer(intern)
            return CustomResponse(response=serializer.data).get_success_response()
            
        interns = UserInternGuildLink.objects.select_related('user').prefetch_related(
            'user__user_role_link_user__role'
        ).order_by('-created_at')
        
        paginated_queryset = CommonUtils.get_paginated_queryset(
            interns, request,
            ['user__full_name', 'guild', 'status'],
            {'created_at': 'created_at', 'status': 'status'}
        )
        
        serializer = ManageInternSerializer(paginated_queryset.get("queryset"), many=True)
        return CustomResponse(
            response={
                "data": serializer.data,
                "pagination": paginated_queryset.get("pagination")
            }
        ).get_success_response()

    @role_required([RoleType.ADMIN.value, RoleType.INTERN.value, RoleType.INTERN_LEAD.value])
    @extend_schema(
        tags=['Dashboard - Intern'],
        description="Onboard a new intern.",
        responses={200: OpenApiResponse(description="Intern onboarded successfully.")},
    )
    def post(self, request):
        user_id = JWTUtils.fetch_user_id(request)
        serializer = ManageInternSerializer(data=request.data, context={'user_id': user_id})
        
        if serializer.is_valid():
            serializer.save()
            return CustomResponse(general_message="Intern onboarded successfully.").get_success_response()
            
        return CustomResponse(response=serializer.errors).get_failure_response()

    @role_required([RoleType.ADMIN.value, RoleType.INTERN.value, RoleType.INTERN_LEAD.value])
    @extend_schema(
        tags=['Dashboard - Intern'],
        description="Update intern details (partial update).",
        responses={200: OpenApiResponse(description="Intern details updated successfully.")},
    )
    def patch(self, request, intern_id):
        user_id = JWTUtils.fetch_user_id(request)
        intern = UserInternGuildLink.objects.filter(id=intern_id).first()
        if not intern:
            return CustomResponse(general_message="Intern not found.").get_failure_response()
            
        old_data = {
            "guild": intern.guild,
            "status": intern.status
        }
            
        serializer = ManageInternSerializer(intern, data=request.data, partial=True, context={'user_id': user_id})
        
        if serializer.is_valid():
            serializer.save()
            
            new_guild = request.data.get("guild")
            if new_guild and new_guild != old_data["guild"]:
                from db.mentor import SystemActionLog
                SystemActionLog.objects.create(
                    action_type=SystemActionLog.ActionType.INTERN_GUILD_REASSIGN.value,
                    actor_user_id=user_id,
                    subject_user_id=intern.user_id,
                    entity_name='user_intern_guild_link',
                    entity_id=intern.id,
                    old_data=old_data,
                    new_data={"guild": new_guild, "status": intern.status}
                )
            
            return CustomResponse(general_message="Intern details updated successfully.").get_success_response()
            
        return CustomResponse(response=serializer.errors).get_failure_response()

    @role_required([RoleType.ADMIN.value, RoleType.INTERN.value, RoleType.INTERN_LEAD.value])
    @extend_schema(
        tags=['Dashboard - Intern'],
        description="Deactivate (soft-delete) an intern.",
        responses={200: OpenApiResponse(description="Intern deactivated successfully.")},
    )
    def delete(self, request, intern_id):
        from django.db import transaction
        from db.user import UserRoleLink
        from utils.types import RoleType
        
        user_id = JWTUtils.fetch_user_id(request)
        intern = UserInternGuildLink.objects.filter(id=intern_id).first()
        if not intern:
            return CustomResponse(general_message="Intern not found.").get_failure_response()
            
        with transaction.atomic():
            # Soft delete / Deactivate
            intern.status = InternGuildStatus.INACTIVE.value
            intern.updated_by_id = user_id
            intern.save()
            
            # Remove "Intern" or "Intern Lead" role
            UserRoleLink.objects.filter(
                user=intern.user, 
                role__title__in=[RoleType.INTERN.value, RoleType.INTERN_LEAD.value]
            ).delete()
        
        return CustomResponse(general_message="Intern deactivated successfully.").get_success_response()

class ManageInternStatusAPI(APIView):
    authentication_classes = [CustomizePermission]

    @role_required([RoleType.ADMIN.value, RoleType.INTERN.value, RoleType.INTERN_LEAD.value])
    @extend_schema(
        tags=['Dashboard - Intern'],
        description="Retrieve intern status statistics.",
        responses={200: OpenApiResponse(description="Intern status counts.")},
    )
    def get(self, request):
        stats = UserInternGuildLink.objects.values('status').annotate(count=Count('id'))
        
        data = {
            InternGuildStatus.ACTIVE.value: 0,
            InternGuildStatus.AT_RISK.value: 0,
            InternGuildStatus.ON_LEAVE.value: 0,
            InternGuildStatus.INACTIVE.value: 0,
        }
        
        for stat in stats:
            data[stat['status']] = stat['count']
            
        data['total_interns'] = UserInternGuildLink.objects.count()
            
        return CustomResponse(response=data).get_success_response()

class ManageInternExportAPI(APIView):
    authentication_classes = [CustomizePermission]

    @role_required([RoleType.ADMIN.value, RoleType.INTERN.value, RoleType.INTERN_LEAD.value])
    @extend_schema(
        tags=['Dashboard - Intern'],
        description="Export interns list as CSV.",
        responses={200: OpenApiResponse(description="CSV attachment of interns.")},
    )
    def get(self, request):
        interns = UserInternGuildLink.objects.all().select_related('user')
        
        # We can reuse the CommonUtils method to generate a CSV response if it exists,
        # but otherwise we can manually build it or use a library.
        # We'll just build a basic CSV string and return it in CustomResponse or as HttpResponse.
        from django.http import HttpResponse
        import csv
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="interns.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Full Name', 'MuID', 'Guild', 'Status', 'Created At'])
        
        for intern in interns:
            writer.writerow([
                intern.user.full_name,
                intern.user.muid,
                intern.guild,
                intern.status,
                intern.created_at.strftime('%Y-%m-%d %H:%M:%S')
            ])
            
        return response


class ManageInternBulkImportTemplateAPIView(APIView):
    authentication_classes = [CustomizePermission]

    @role_required([RoleType.ADMIN.value])
    @extend_schema(
        tags=['Dashboard - Intern'],
        description="Download the Excel template for bulk intern import. Columns: muid, guild.",
        responses={200: OpenApiResponse(description="Excel file download.")},
    )
    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Interns"
        ws.append(['muid', 'guild'])

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = FileResponse(
            output,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="intern_bulk_import_template.xlsx"'
        return response


class ManageInternBulkImportAPI(APIView):
    authentication_classes = [CustomizePermission]
    parser_classes = [MultiPartParser, FormParser]

    @role_required([RoleType.ADMIN.value])
    @extend_schema(
        tags=['Dashboard - Intern'],
        description=(
            "Bulk assign users as interns by uploading an Excel (.xlsx) file. "
            "Required columns: 'muid', 'guild'. Role defaults to 'Intern'. "
            "Rows where the user is already an intern are skipped."
        ),
        responses={200: OpenApiResponse(description="Bulk import result.")},
    )
    def post(self, request):
        actor_user_id = JWTUtils.fetch_user_id(request)

        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            return CustomResponse(
                general_message="No file uploaded. Please attach an Excel file with key 'excel_file'."
            ).get_failure_response()

        # Validate file extension
        if not excel_file.name.endswith('.xlsx'):
            return CustomResponse(
                general_message="Invalid file type. Only .xlsx files are supported."
            ).get_failure_response()

        try:
            wb = openpyxl.load_workbook(excel_file)
        except Exception:
            return CustomResponse(
                general_message="Could not read the uploaded file. Please ensure it is a valid .xlsx file."
            ).get_failure_response()

        ws = wb.active
        headers = [cell.value for cell in ws[1]] if ws.max_row > 0 else []

        # Validate required headers
        required_headers = ['muid', 'guild']
        missing_headers = [h for h in required_headers if h not in headers]
        if missing_headers:
            return CustomResponse(
                general_message=f"Missing required column(s): {', '.join(missing_headers)}. "
                                f"Expected headers: {required_headers}."
            ).get_failure_response()

        muid_idx = headers.index('muid')
        guild_idx = headers.index('guild')

        success_count = 0
        failed_rows = []

        # Fetch the Intern role once
        intern_role = Role.objects.filter(title=RoleType.INTERN.value).first()
        if not intern_role:
            return CustomResponse(
                general_message=f"Role '{RoleType.INTERN.value}' does not exist in the system. "
                                f"Please contact a system administrator."
            ).get_failure_response()

        # Track muids processed in this file to handle in-file duplicates
        seen_muids = set()

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            muid = row[muid_idx]
            guild = row[guild_idx]

            # Skip entirely empty rows
            if muid is None and guild is None:
                continue

            # Coerce to string and strip whitespace if present
            muid = str(muid).strip() if muid is not None else None
            guild = str(guild).strip() if guild is not None else None

            # --- Edge case: missing muid or guild value ---
            if not muid or not guild:
                failed_rows.append({
                    "row": row_num,
                    "muid": muid or "(empty)",
                    "reason": "Both 'muid' and 'guild' values are required."
                })
                continue

            # --- Edge case: duplicate muid within the same file ---
            if muid in seen_muids:
                failed_rows.append({
                    "row": row_num,
                    "muid": muid,
                    "reason": "Duplicate muid in file — already processed earlier in this upload."
                })
                continue

            seen_muids.add(muid)

            try:
                with transaction.atomic():
                    # --- Edge case: user not found ---
                    user = User.objects.filter(muid=muid).first()
                    if not user:
                        failed_rows.append({
                            "row": row_num,
                            "muid": muid,
                            "reason": "No user found with this muid."
                        })
                        continue

                    # --- Edge case: user is already an intern ---
                    if UserInternGuildLink.objects.filter(user=user).exists():
                        failed_rows.append({
                            "row": row_num,
                            "muid": muid,
                            "reason": "User is already onboarded as an intern."
                        })
                        continue

                    # Create the intern guild link
                    UserInternGuildLink.objects.create(
                        id=str(uuid.uuid4()),
                        user=user,
                        guild=guild,
                        status=InternGuildStatus.ACTIVE.value,
                        created_by_id=actor_user_id,
                        updated_by_id=actor_user_id,
                    )

                    # Assign 'Intern' role (idempotent via get_or_create)
                    UserRoleLink.objects.get_or_create(
                        user=user,
                        role=intern_role,
                        defaults={
                            'verified': True,
                            'is_active': True,
                            'created_by_id': actor_user_id
                        }
                    )

                    success_count += 1

            except Exception as e:
                failed_rows.append({
                    "row": row_num,
                    "muid": muid,
                    "reason": f"Unexpected error: {str(e)}"
                })

        return CustomResponse(
            general_message="Bulk intern import completed.",
            response={
                "success_count": success_count,
                "failed_count": len(failed_rows),
                "failed_rows": failed_rows,
            }
        ).get_success_response()
