import uuid
from django.db import IntegrityError
from rest_framework.views import APIView

from db.user import Role, User, UserRoleLink
from utils.permission import CustomizePermission, role_required, JWTUtils
from utils.response import CustomResponse
from utils.types import RoleType, WebHookActions, WebHookCategory
from utils.utils import CommonUtils, DiscordWebhooks, ImportCSV
from . import dash_roles_serializer

from openpyxl import load_workbook
from tempfile import NamedTemporaryFile
from io import BytesIO
from django.http import FileResponse
from django.db.models import Q


class RoleAPI(APIView):
    authentication_classes = [CustomizePermission]

    @role_required([RoleType.ADMIN.value])
    def get(self, request):
        roles_queryset = Role.objects.all()

        queryset = CommonUtils.get_paginated_queryset(
            roles_queryset,
            request,
            [
                "id",
                "title",
                "description",
                "updated_by__first_name",
                "updated_by__last_name",
                "created_by__first_name",
                "created_by__last_name",
            ],
            {
                "title": "title",
                "description": "description",
                "members": "userrolelink",
                "updated_by": "updated_by__first_name",
                "created_by": "created_by__first_name",
                "updated_at": "updated_at",
                "created_at": "created_at",
            },
        )

        serializer = dash_roles_serializer.RoleDashboardSerializer(
            queryset.get("queryset"), many=True
        )

        return CustomResponse().paginated_response(
            data=serializer.data, pagination=queryset.get("pagination")
        )

    @role_required([RoleType.ADMIN.value])
    def patch(self, request, roles_id):
        role = Role.objects.get(id=roles_id)
        old_name = role.title

        serializer = dash_roles_serializer.RoleDashboardSerializer(
            role, data=request.data, partial=True, context={"request": request}
        )

        if not serializer.is_valid():
            return CustomResponse(
                general_message=serializer.errors
            ).get_failure_response()

        try:
            serializer.save()
            newname = role.title

            DiscordWebhooks.general_updates(
                WebHookCategory.ROLE.value, WebHookActions.EDIT.value, newname, old_name
            )

            return CustomResponse(
                response={"data": serializer.data}
            ).get_success_response()

        except IntegrityError as e:
            return CustomResponse(
                general_message="Database integrity error",
            ).get_failure_response()

    @role_required([RoleType.ADMIN.value])
    def delete(self, request, roles_id):
        role = Role.objects.get(id=roles_id)
        role.delete()

        DiscordWebhooks.general_updates(
            WebHookCategory.ROLE.value, WebHookActions.DELETE.value, role.title
        )
        return CustomResponse(
            general_message="Role deleted successfully"
        ).get_success_response()

    @role_required([RoleType.ADMIN.value])
    def post(self, request):
        serializer = dash_roles_serializer.RoleDashboardSerializer(
            data=request.data, partial=True, context={"request": request}
        )

        if serializer.is_valid():
            serializer.save()

            DiscordWebhooks.general_updates(
                WebHookCategory.ROLE.value,
                WebHookActions.CREATE.value,
                request.data.get("title"),
            )

            return CustomResponse(
                response={"data": serializer.data}
            ).get_success_response()

        return CustomResponse(general_message=serializer.errors).get_failure_response()


class RoleManagementCSV(APIView):
    authentication_classes = [CustomizePermission]

    @role_required([RoleType.ADMIN.value])
    def get(self, request):
        role = Role.objects.all()

        role_serializer_data = dash_roles_serializer.RoleDashboardSerializer(
            role, many=True
        ).data
        return CommonUtils.generate_csv(role_serializer_data, "Roles")


class UserRoleSearchAPI(APIView):
    authentication_classes = [CustomizePermission]

    @role_required([RoleType.ADMIN.value])
    def get(self, request, role_id):
        user = User.objects.filter(user_role_link_user__role_id=role_id).distinct()

        paginated_queryset = CommonUtils.get_paginated_queryset(
            user,
            request,
            ["muid", "first_name", "last_name"],
            {"muid": "muid", "first_name": "first_name", "last_name": "last_name"},
        )

        serializer = dash_roles_serializer.UserRoleSearchSerializer(
            paginated_queryset.get("queryset"), many=True
        ).data

        return CustomResponse(
            response={
                "data": serializer,
                "pagination": paginated_queryset.get("pagination"),
            }
        ).get_success_response()


class UserRoleLinkManagement(APIView):
    authentication_classes = [CustomizePermission]
    """
    This API is creates an interface to help manage the user and role link
    by providing support for 
    - Listing all users with the given role
    - Giving a lot of users a specific role
    """

    @role_required([RoleType.ADMIN.value])
    def get(self, request, role_id):
        """
        Lists all the users with a given role
        """
        users = (
            User.objects.filter(user_role_link_user__role__pk=role_id).distinct().all()
        )
        serialized_users = dash_roles_serializer.UserRoleLinkManagementSerializer(
            users, many=True
        )
        return CustomResponse(response=serialized_users.data).get_success_response()

    @role_required([RoleType.ADMIN.value])
    def put(self, request, role_id):
        """
        Lists all the users without a given role;
        used to assign roles
        """
        users = (
            User.objects.filter(~Q(user_role_link_user__role__pk=role_id))
            .distinct()
            .all()
        )
        serialized_users = dash_roles_serializer.UserRoleLinkManagementSerializer(
            users, many=True
        )
        return CustomResponse(response=serialized_users.data).get_success_response()

    @role_required([RoleType.ADMIN.value])
    def post(self, request, role_id):
        """
        Assigns a large bunch of users a certain role
        """
        request_data = request.data.copy()
        request_data["role"] = role_id
        request_data["created_by"] = JWTUtils.fetch_user_id(request)
        serialized_users = dash_roles_serializer.RoleAssignmentSerializer(
            data=request_data
        )
        if serialized_users.is_valid():
            users, role = serialized_users.save()
            return CustomResponse(
                general_message=f"Successfully gave {len(users)} users '{role.title}' role"
            ).get_success_response()
        return CustomResponse(response=serialized_users.errors).get_failure_response()

    @role_required([RoleType.ADMIN.value])
    def patch(self, request, role_id):
        """
        Removes a role from a large bunch of users
        """
        try:
            role = Role.objects.get(pk=role_id)
            user_role_links = UserRoleLink.objects.filter(
                user__pk__in=request.data.get("users"),
                role=role,
            )
            number = user_role_links.count()

            DiscordWebhooks.general_updates(
                WebHookCategory.BULK_ROLE.value,
                WebHookActions.DELETE.value,
                role.title,
                ",".join(list(user_role_links.values_list("user_id", flat=True))),
            )

            user_role_links.delete()

            return CustomResponse(
                general_message=(
                    f"Successfully removed the '{role.title}' role from {number} users"
                )
            ).get_success_response()
        except Role.DoesNotExist as e:
            return CustomResponse(general_message=str(e)).get_failure_response()


class UserRole(APIView):
    authentication_classes = [CustomizePermission]

    @role_required([RoleType.ADMIN.value])
    def post(self, request):
        serializer = dash_roles_serializer.UserRoleCreateSerializer(
            data=request.data, context={"request": request}
        )

        if not serializer.is_valid():
            return CustomResponse(
                general_message=serializer.errors
            ).get_failure_response()

        serializer.save()

        DiscordWebhooks.general_updates(
            WebHookCategory.USER_ROLE.value,
            WebHookActions.UPDATE.value,
            request.data.get("user_id"),
        )
        return CustomResponse(
            general_message="Role Added Successfully"
        ).get_success_response()

    @role_required([RoleType.ADMIN.value])
    def delete(self, request):
        serializer = dash_roles_serializer.UserRoleCreateSerializer(
            data=request.data, context={"request": request}
        )

        if not serializer.is_valid():
            return CustomResponse(
                general_message=serializer.errors
            ).get_failure_response()

        user_id = request.data.get("user_id")
        role_id = request.data.get("role_id")

        user_role_link = UserRoleLink.objects.get(role_id=role_id, user_id=user_id)

        user_role_link.delete()

        DiscordWebhooks.general_updates(
            WebHookCategory.USER_ROLE.value,
            WebHookActions.DELETE.value,
            user_role_link.id,
        )
        return CustomResponse(
            general_message="User Role deleted successfully"
        ).get_success_response()


class RoleBaseTemplateAPI(APIView):
    authentication_classes = [CustomizePermission]

    def get(self, request):
        wb = load_workbook("./excel-templates/role_base_template.xlsx")
        ws = wb["Data Definitions"]

        roles = Role.objects.all().values_list("title", flat=True)
        data = {"role": roles}
        # Write data column-wise
        for col_num, (col_name, col_values) in enumerate(data.items(), start=1):
            for row, value in enumerate(col_values, start=2):
                ws.cell(row=row, column=col_num, value=value)

        # Save the file
        with NamedTemporaryFile() as tmp:
            tmp.close()  # with statement opened tmp, close it so wb.save can open it
            wb.save(tmp.name)
            with open(tmp.name, "rb") as f:
                f.seek(0)
                new_file_object = f.read()
        return FileResponse(
            BytesIO(new_file_object),
            as_attachment=True,
            filename="role_base_template.xlsx",
        )


class UserRoleBulkAssignAPI(APIView):
    authentication_classes = [CustomizePermission]

    @role_required([RoleType.ADMIN.value])
    def post(self, request):
        try:
            file_obj = request.FILES["user_roles_list"]
        except KeyError:
            return CustomResponse(
                general_message="File not found."
            ).get_failure_response()

        excel_data = ImportCSV()
        excel_data = excel_data.read_excel_file(file_obj)

        if not excel_data:
            return CustomResponse(
                general_message="Empty csv file."
            ).get_failure_response()

        temp_headers = ["muid", "role"]
        first_entry = excel_data[0]
        for key in temp_headers:
            if key not in first_entry:
                return CustomResponse(
                    general_message=f"{key} does not exist in the file."
                ).get_failure_response()

        excel_data = [row for row in excel_data if any(row.values())]
        valid_rows = []
        error_rows = []
        users_to_fetch = set()
        roles_to_fetch = set()
        user_role_link_to_check = set()

        for row in excel_data[1:]:
            keys_to_keep = ["muid", "role"]
            row_keys = list(row.keys())
            # Remove columns other than "muid" and "role"
            for key in row_keys:
                if key not in keys_to_keep:
                    del row[key]

        for row in excel_data[1:]:
            user = row.get("muid")
            role = row.get("role")
            users_to_fetch.add(user)
            roles_to_fetch.add(role)
            if (user, role) in user_role_link_to_check:
                row["error"] = "Duplicate entry"
                error_rows.append(row)
                excel_data.remove(row)
            else:
                user_role_link_to_check.add((user, role))

        users = User.objects.filter(muid__in=users_to_fetch).values(
            "id",
            "muid",
        )
        roles = Role.objects.filter(title__in=roles_to_fetch).values(
            "id",
            "title",
        )
        existing_user_role_links = list(
            UserRoleLink.objects.filter(
                user__muid__in=users_to_fetch, role__title__in=roles_to_fetch
            ).values_list("user__muid", "role__title")
        )
        users_dict = {user["muid"]: user["id"] for user in users}
        roles_dict = {role["title"]: role["id"] for role in roles}
        users_by_role = {role_title: [] for role_title in roles_dict.keys()}

        for row in excel_data[1:]:
            user = row.pop("muid")
            role = row.pop("role")

            user_id = users_dict.get(user)
            role_id = roles_dict.get(role)
            if not user_id:
                row["muid"] = user
                row["role"] = role
                row["error"] = f"Invalid user muid: {user}"
                error_rows.append(row)
            elif not role_id:
                row["muid"] = user
                row["role"] = role
                row["error"] = f"Invalid role: {role}"
                error_rows.append(row)
            elif (user, role) in existing_user_role_links:
                row["muid"] = user
                row["role"] = role
                row["error"] = f"User {user} already has role {role}"
                error_rows.append(row)
            else:
                users_by_role[role].append(user_id)
                request_user_id = JWTUtils.fetch_user_id(request)
                row["id"] = str(uuid.uuid4())
                row["user_id"] = user_id
                row["role_id"] = role_id
                row["verified"] = True
                row["created_by_id"] = request_user_id
                valid_rows.append(row)

        users_by_role = {
            role_title: users for role_title, users in users_by_role.items() if users
        }
        user_roles_serializer = dash_roles_serializer.UserRoleBulkAssignSerializer(
            data=valid_rows, many=True
        )
        success_data = []
        if user_roles_serializer.is_valid():
            user_roles_serializer.save()
            for user_role_data in user_roles_serializer.data:
                success_data.append(
                    {
                        "user": user_role_data.get("user_id", ""),
                        "role": user_role_data.get("role_id", ""),
                    }
                )
            for role, user_set in users_by_role.items():
                DiscordWebhooks.general_updates(
                    WebHookCategory.BULK_ROLE.value,
                    WebHookActions.UPDATE.value,
                    role,
                    ",".join(user_set),
                )
        else:
            error_rows.append(user_roles_serializer.errors)

        return CustomResponse(
            response={"Success": success_data, "Failed": error_rows}
        ).get_success_response()
