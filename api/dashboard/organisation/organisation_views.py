import uuid
from django.db.models import F, Prefetch, Sum
from rest_framework.views import APIView

from openpyxl import load_workbook
from tempfile import NamedTemporaryFile
from io import BytesIO
from django.http import FileResponse

from db.organization import (
    Department,
    OrgAffiliation,
    Organization,
    UserOrganizationLink,
    District,
)
from utils.permission import CustomizePermission, JWTUtils, role_required
from utils.response import CustomResponse
from utils.types import OrganizationType, RoleType, WebHookActions, WebHookCategory
from utils.utils import CommonUtils, DiscordWebhooks, ImportCSV

from .serializers import (
    AffiliationCreateUpdateSerializer,
    AffiliationSerializer,
    DepartmentSerializer,
    InstitutionCreateUpdateSerializer,
    InstitutionSerializer,
    InstitutionPrefillSerializer,
    OrganizationMergerSerializer, OrganizationKarmaTypeGetPostPatchDeleteSerializer,
    OrganizationKarmaLogGetPostPatchDeleteSerializer,
    OrganizationImportSerializer
)


class InstitutionPostUpdateDeleteAPI(APIView):
    authentication_classes = [CustomizePermission]

    @role_required([RoleType.ADMIN.value])
    def post(self, request):
        user_id = JWTUtils.fetch_user_id(request)

        serializer = InstitutionCreateUpdateSerializer(
            data=request.data, context={"user_id": user_id}
        )

        if serializer.is_valid():
            serializer.save()

            if request.data.get("org_type") == OrganizationType.COMMUNITY.value:
                DiscordWebhooks.general_updates(
                    WebHookCategory.COMMUNITY.value,
                    WebHookActions.CREATE.value,
                    request.data.get("title"),
                )

            return CustomResponse(
                general_message="Organisation added successfully"
            ).get_success_response()

        return CustomResponse(general_message=serializer.errors).get_failure_response()

    @role_required([RoleType.ADMIN.value])
    def put(self, request, org_code):
        user_id = JWTUtils.fetch_user_id(request)

        organization = Organization.objects.filter(code=org_code).first()

        if organization is None:
            return CustomResponse(
                general_message="Invalid organization code"
            ).get_failure_response()

        old_title = organization.title
        old_type = organization.org_type

        serializer = InstitutionCreateUpdateSerializer(
            organization, data=request.data, context={"user_id": user_id}
        )

        if serializer.is_valid():
            serializer.save()

            if (
                request.data.get("title") != old_title
                and old_type == OrganizationType.COMMUNITY.value
            ):
                DiscordWebhooks.general_updates(
                    WebHookCategory.COMMUNITY.value,
                    WebHookActions.EDIT.value,
                    request.data.get("title"),
                    old_title,
                )

            if (
                request.data.get("orgType") != OrganizationType.COMMUNITY.value
                and old_type == OrganizationType.COMMUNITY.value
            ):
                DiscordWebhooks.general_updates(
                    WebHookCategory.COMMUNITY.value,
                    WebHookActions.DELETE.value,
                    old_title,
                )

            if (
                old_type != OrganizationType.COMMUNITY.value
                and request.data.get("orgType") == OrganizationType.COMMUNITY.value
            ):
                title = request.data.get("title") or old_title
                DiscordWebhooks.general_updates(
                    WebHookCategory.COMMUNITY.value, WebHookActions.CREATE.value, title
                )

            return CustomResponse(
                general_message="Organization edited successfully"
            ).get_success_response()

        return CustomResponse(message=serializer.errors).get_failure_response()

    @role_required([RoleType.ADMIN.value])
    def delete(self, request, org_code):
        if not (organisation := Organization.objects.filter(code=org_code).first()):
            return CustomResponse(
                general_message=f"Org with code '{org_code}', does not exist"
            ).get_failure_response()

        organisation.delete()
        org_type = organisation.org_type

        if org_type == OrganizationType.COMMUNITY.value:
            DiscordWebhooks.general_updates(
                WebHookCategory.COMMUNITY.value,
                WebHookActions.DELETE.value,
                organisation.title,
            )

        return CustomResponse(
            general_message="Organization deleted successfully"
        ).get_success_response()


class InstitutionAPI(APIView):
    def get(self, request, org_type, district_id=None):
        if district_id:
            organisations = Organization.objects.filter(
                org_type=org_type, district_id=district_id
            )
        else:
            organisations = Organization.objects.filter(org_type=org_type)

        org_queryset = organisations.select_related(
            "affiliation",
            # 'district',
            # "district__zone__state__country",
            # "district__zone__state",
            # "district__zone",
            # "district",
        ).prefetch_related(
            Prefetch(
                "user_organization_link_org",
                queryset=UserOrganizationLink.objects.filter(
                    verified=True
                ).select_related("user"),
            )
        )

        paginated_queryset = CommonUtils.get_paginated_queryset(
            org_queryset,
            request,
            [
                "title",
                "code",
                "affiliation__title",
                "district__name",
                "district__zone__name",
                "district__zone__state__name",
                "district__zone__state__country__name",
            ],
            {
                "title": "title",
                "code": "code",
                "affiliation": "affiliation__title",
                "district": "district__name",
                "zone": "district__zone__name",
                "state": "district__zone__state__name",
                "country": "district__zone__state__country__name",
            },
        )

        serializer = InstitutionSerializer(
            paginated_queryset.get("queryset"), many=True
        )

        return CustomResponse(
            response={
                "data": serializer.data,
                "pagination": paginated_queryset.get("pagination"),
            }
        ).get_success_response()


class InstitutionCSVAPI(APIView):
    authentication_classes = [CustomizePermission]

    @role_required([RoleType.ADMIN.value])
    def get(self, request, org_type):
        organizations = (
            Organization.objects.filter(org_type=org_type)
            .select_related(
                "affiliation",
                # "district__zone__state__country",
                # "district__zone__state",
                # "district__zone",
                # "district",
            )
            .prefetch_related(
                Prefetch(
                    "user_organization_link_org",
                    queryset=UserOrganizationLink.objects.filter(
                        verified=True
                    ).select_related("user"),
                )
            )
        )

        serializer = InstitutionSerializer(organizations, many=True).data

        return CommonUtils.generate_csv(serializer, f"{org_type} data")


class InstitutionDetailsAPI(APIView):
    @role_required(
        [
            RoleType.ADMIN.value,
        ]
    )
    def get(self, request, org_code):
        organization = (
            Organization.objects.filter(code=org_code)
            .values(
                "id",
                "title",
                "code",
                affiliation_uuid=F("affiliation__id"),
                affiliation_name=F("affiliation__title"),
                district_uuid=F("district__id"),
                district_name=F("district__name"),
                zone_uuid=F("district__zone__id"),
                zone_name=F("district__zone__name"),
                state_uuid=F("district__zone__state__id"),
                state_name=F("district__zone__state__name"),
                country_uuid=F("district__zone__state__country__id"),
                country_name=F("district__zone__state__country__name"),
            )
            .annotate(karma=Sum("user_organization_link_org__user__wallet_user__karma"))
            .order_by("-karma")
        )

        return CustomResponse(response=organization).get_success_response()


class GetInstitutionsAPI(APIView):
    def get(self, request, org_type, district_id=None):
        if district_id:
            organisations = Organization.objects.filter(
                org_type=org_type, district_id=district_id
            )
        else:
            organisations = Organization.objects.filter(org_type=org_type)

        paginated_organisations = CommonUtils.get_paginated_queryset(
            organisations, request, ["title", "code"]
        )

        organisation_serializer = InstitutionSerializer(
            paginated_organisations.get("queryset"), many=True
        )
        return CustomResponse().paginated_response(
            data=organisation_serializer.data,
            pagination=paginated_organisations.get("pagination"),
        )


class AffiliationGetPostUpdateDeleteAPI(APIView):
    authentication_classes = [CustomizePermission]

    @role_required([RoleType.ADMIN.value])
    def get(self, request):
        affiliation = OrgAffiliation.objects.all()
        paginated_queryset = CommonUtils.get_paginated_queryset(
            affiliation, request, ["id", "title"]
        )
        serializer = AffiliationSerializer(
            paginated_queryset.get("queryset"), many=True
        )
        return CustomResponse().paginated_response(
            data=serializer.data, pagination=paginated_queryset.get("pagination")
        )

    @role_required([RoleType.ADMIN.value])
    def post(self, request):
        user_id = JWTUtils.fetch_user_id(request)

        serializer = AffiliationCreateUpdateSerializer(
            data=request.data,
            context={
                "user_id": user_id,
            },
        )

        if serializer.is_valid():
            serializer.save()

            return CustomResponse(
                general_message=f"{request.data.get('title')} added successfully"
            ).get_success_response()

        return CustomResponse(general_message=serializer.errors).get_failure_response()

    @role_required([RoleType.ADMIN.value])
    def put(self, request, affiliation_id):
        user_id = JWTUtils.fetch_user_id(request)

        affiliation = OrgAffiliation.objects.filter(id=affiliation_id).first()

        if affiliation is None:
            return CustomResponse(
                general_message="Invalid affiliation"
            ).get_failure_response()

        serializer = AffiliationCreateUpdateSerializer(
            affiliation, data=request.data, context={"user_id": user_id}
        )

        if serializer.is_valid():
            serializer.save()

            return CustomResponse(
                general_message=f"{affiliation.title} Edited Successfully"
            ).get_success_response()

        return CustomResponse(message=serializer.errors).get_failure_response()

    @role_required([RoleType.ADMIN.value])
    def delete(self, request, affiliation_id):
        affiliation = OrgAffiliation.objects.filter(id=affiliation_id).first()

        if affiliation is None:
            return CustomResponse(
                general_message="Invalid affiliation id"
            ).get_failure_response()

        affiliation.delete()

        return CustomResponse(
            general_message=f"{affiliation.title} Deleted Successfully"
        ).get_success_response()


class DepartmentAPI(APIView):
    authentication_classes = [CustomizePermission]

    @role_required([RoleType.ADMIN.value])
    def get(self, request, dept_id=None):
        if dept_id:
            departments = Department.objects.filter(id=dept_id)
        else:
            departments = Department.objects.all()

        paginated_queryset = CommonUtils.get_paginated_queryset(
            departments, request, ["title"], {"title": "title"}
        )

        serializer = DepartmentSerializer(paginated_queryset.get("queryset"), many=True)

        return CustomResponse().paginated_response(
            data=serializer.data, pagination=paginated_queryset.get("pagination")
        )

    @role_required([RoleType.ADMIN.value])
    def post(self, request):
        serializer = DepartmentSerializer(
            data=request.data, context={"request": request}
        )

        if serializer.is_valid():
            serializer.save()

            return CustomResponse(
                general_message=f"{request.data.get('title')} created successfully"
            ).get_success_response()

        return CustomResponse(response=serializer.errors).get_failure_response()

    @role_required([RoleType.ADMIN.value])
    def put(self, request, department_id):
        department = Department.objects.get(id=department_id)

        serializer = DepartmentSerializer(
            department, data=request.data, context={"request": request}
        )

        if serializer.is_valid():
            serializer.save()

            return CustomResponse(
                general_message=f"{department.title} updated successfully"
            ).get_success_response()

        return CustomResponse(response=serializer.errors).get_failure_response()

    @role_required([RoleType.ADMIN.value])
    def delete(self, request, department_id):
        department = Department.objects.get(id=department_id)

        department.delete()
        return CustomResponse(
            general_message=f"{department.title} deleted successfully"
        ).get_success_response()


class AffiliationListAPI(APIView):
    @role_required([RoleType.ADMIN.value])
    def get(self, request):
        affiliation = OrgAffiliation.objects.all().values("id", "title")

        return CustomResponse(response=affiliation).get_success_response()


class InstitutionPrefillAPI(APIView):
    @role_required(
        [
            RoleType.ADMIN.value,
        ]
    )
    def get(self, request, org_code):
        organization = Organization.objects.filter(code=org_code).first()

        serializer = InstitutionPrefillSerializer(organization, many=False).data

        return CustomResponse(response=serializer).get_success_response()


class OrganizationMergerView(APIView):
    permission_classes = [CustomizePermission]

    @role_required([RoleType.ADMIN.value])
    def get(self, request, organisation_id):
        try:
            destination = Organization.objects.get(pk=organisation_id)
            serializer = OrganizationMergerSerializer(destination, data=request.data)
            if not serializer.is_valid():
                return CustomResponse(
                    general_message=serializer.errors
                ).get_failure_response()

            return CustomResponse(response=serializer.data).get_success_response()

        except Organization.DoesNotExist:
            return CustomResponse(
                general_message="An organization with the given id doesn't exist"
            ).get_failure_response()

    @role_required([RoleType.ADMIN.value])
    def patch(self, request, organisation_id):
        try:
            destination = Organization.objects.get(pk=organisation_id)
            serializer = OrganizationMergerSerializer(destination, data=request.data)
            if not serializer.is_valid():
                return CustomResponse(
                    general_message=serializer.errors
                ).get_failure_response()
            serializer.save()

            return CustomResponse(
                general_message=f"Organizations merged successfully into {destination.title}."
            ).get_success_response()

        except Organization.DoesNotExist:
            return CustomResponse(
                general_message="An organization with the given id doesn't exist"
            ).get_failure_response()


class OrganizationKarmaTypeGetPostPatchDeleteAPI(APIView):
    def post(self, request):
        user_id = JWTUtils.fetch_user_id(request)

        serializer = OrganizationKarmaTypeGetPostPatchDeleteSerializer(
            data=request.data,
            context={
                'user_id': user_id
            }
        )
        if serializer.is_valid():
            serializer.save()

            return CustomResponse(
                "Organization karma type created successfully"
            ).get_success_response()

        return CustomResponse(
            serializer.errors
        ).get_failure_response()


class OrganizationKarmaLogGetPostPatchDeleteAPI(APIView):
    def post(self, request):
        user_id = JWTUtils.fetch_user_id(request)

        serializer = OrganizationKarmaLogGetPostPatchDeleteSerializer(
            data=request.data,
            context={
                'user_id': user_id
            }
        )

        if serializer.is_valid():
            serializer.save()

            return CustomResponse(
                "Organization karma Log created successfully"
            ).get_success_response()

        return CustomResponse(
            serializer.errors
        ).get_failure_response()

class OrganisationBaseTemplateAPI(APIView):
    authentication_classes = [CustomizePermission]
    
    def get(self, request):
        wb = load_workbook('./api/dashboard/organisation/assets/organisation_base_template.xlsx')
        ws = wb['Data Definitions']
        affiliations = OrgAffiliation.objects.all().values_list('title', flat=True)
        districts = District.objects.all().values_list('name', flat=True)
        org_types = OrganizationType.get_all_values()

        data = {
            'org_type': org_types,
            'affiliation': affiliations,
            'district': districts,
        }
        # Write data column-wise
        for col_num, (col_name, col_values) in enumerate(data.items(), start=1):
            for row, value in enumerate(col_values, start=2):
                ws.cell(row=row, column=col_num, value=value)
        # Save the file
        with NamedTemporaryFile() as tmp:
            tmp.close() # with statement opened tmp, close it so wb.save can open it
            wb.save(tmp.name)
            with open(tmp.name, 'rb') as f:
                f.seek(0)
                new_file_object = f.read()
        return FileResponse(BytesIO(new_file_object), as_attachment=True, filename='organisation_base_template.xlsx')
    
class OrganisationImportAPI(APIView):
    authentication_classes = [CustomizePermission]

    @role_required([RoleType.ADMIN.value])
    def post(self, request):
        try:
            file_obj = request.FILES["organisation_list"]
        except KeyError:
            return CustomResponse(
                general_message="File not found."
            ).get_failure_response()

        excel_data = ImportCSV()
        excel_data = excel_data.read_excel_file(file_obj)

        if not excel_data:
            return CustomResponse(
                general_message="Empty csv file."
            ).get_failure_response()

        temp_headers = [
            "title",
            "code",
            "org_type",
            "affiliation",
            "district"
        ]
        first_entry = excel_data[0]
        for key in temp_headers:
            if key not in first_entry:
                return CustomResponse(
                    general_message=f"{key} does not exist in the file."
                ).get_failure_response()

        excel_data = [row for row in excel_data if any(row.values())]
        valid_rows = []
        error_rows = []

        title_excel = set()
        code_excel = set()
        title_db = Organization.objects.values_list("title", flat=True)
        code_db = Organization.objects.values_list("code", flat=True)

        affiliations_to_fetch = set()
        districts_to_fetch = set()

        for row in excel_data[1:]:
            title = row.get("title")
            if not title:
                row["error"] = "Missing title."
                error_rows.append(row)
                excel_data.remove(row)
                continue
            elif title in title_excel:
                row["error"] = f"Duplicate title in excel: {title}"
                error_rows.append(row)
                excel_data.remove(row)
                continue
            elif title in title_db:
                row["error"] = f"Duplicate title in database: {title}"
                error_rows.append(row)
                excel_data.remove(row)
                continue
            else:
                title_excel.add(title)

            code = row.get("code")
            if not code:
                row["error"] = "Missing code."
                error_rows.append(row)
                excel_data.remove(row)
                continue
            elif code in code_excel:
                row["error"] = f"Duplicate code in excel: {code}"
                error_rows.append(row)
                excel_data.remove(row)
                continue
            elif code in code_db:
                row["error"] = f"Duplicate code in database: {code}"
                error_rows.append(row)
                excel_data.remove(row)
                continue
            else:
                code_excel.add(code)

            affiliation = row.get("affiliation")
            district = row.get("district")

            affiliations_to_fetch.add(affiliation)
            districts_to_fetch.add(district)

        affiliations = OrgAffiliation.objects.filter(
            title__in=affiliations_to_fetch
        ).values(
            "id",
            "title"
        )

        districts = District.objects.filter(
            name__in=districts_to_fetch
        ).values(
            "id",
            "name"
        )

        affiliations_dict = {affiliation["title"]: affiliation["id"] for affiliation in affiliations}
        districts_dict = {district["name"]: district["id"] for district in districts}
        org_types = OrganizationType.get_all_values()

        for row in excel_data[1:]:
            affiliation = row.pop("affiliation")
            district = row.pop("district")

            affiliation_id = affiliations_dict.get(affiliation) if affiliation is not None else None
            district_id = districts_dict.get(district)
            org_type = row.get("org_type")

            if affiliation and not affiliation_id:
                row["error"] = f"Invalid affiliation: {affiliation}"
                error_rows.append(row)
                print(row)
            elif not district_id:
                row["error"] = f"Invalid district: {district}"
                error_rows.append(row)
                print(row)
            elif org_type not in org_types:
                row["error"] = f"Invalid org_type: {org_type}"
                error_rows.append(row)
                print(row)
            else:
                user_id = JWTUtils.fetch_user_id(request)
                row["id"] = str(uuid.uuid4())
                row["updated_by_id"] = user_id
                row["created_by_id"] = user_id
                row["affiliation_id"] = affiliation_id
                row["district_id"] = district_id
                valid_rows.append(row)

        organization_list_serializer = OrganizationImportSerializer(data=valid_rows, many=True)
        if organization_list_serializer.is_valid():
            organization_list_serializer.save()
        else:
            error_rows.append(organization_list_serializer.errors)
            
        # task_list_serializer = TaskImportSerializer(data=valid_rows, many=True)
        # success_data = []
        # if task_list_serializer.is_valid():
        #     task_list_serializer.save()
        #     for task_data in task_list_serializer.data:
        #             success_data.append({
        #             'hashtag': task_data.get('hashtag', ''),
        #             'title': task_data.get('title', ''),
        #             'description': task_data.get('description', ''),
        #             'karma': task_data.get('karma', ''),
        #             'usage_count': task_data.get('usage_count', ''),
        #             'variable_karma': task_data.get('variable_karma', ''),
        #             'level': task_data.get('level_id', ''),
        #             'channel': task_data.get('channel_id', ''),
        #             'type': task_data.get('type_id', ''),
        #             'ig': task_data.get('ig_id', ''),
        #             'org': task_data.get('org_id', ''),
        #             'event': task_data.get('event', ''),
        #         })
        # else:
        #     error_rows.append(task_list_serializer.errors)

        return CustomResponse(
            response={"Success": organization_list_serializer.data, "Failed": error_rows}
        ).get_success_response()